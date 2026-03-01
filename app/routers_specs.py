from fastapi import APIRouter, Depends, Body, HTTPException, Query, UploadFile, File, Response
from fastapi.responses import StreamingResponse
from io import BytesIO
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Border, Side, Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from collections import defaultdict
from starlette.responses import StreamingResponse
from typing import Any, Dict, Optional, List
from datetime import datetime
from bson import ObjectId
from .audit import log_user_action, log_project_action, make_diff
from . import auth
from .db import db
import re
from motor.motor_asyncio import AsyncIOMotorGridFSBucket
from gridfs.errors import NoFile
from urllib.parse import quote

router = APIRouter()

# хелперы


def _oid(x):
    if isinstance(x, ObjectId):
        return x
    try:
        return ObjectId(x)
    except Exception:
        return x


def _norm(v: Any) -> Any:
    """Преобразовать ObjectId/даты в json-дружественный вид, рекурсивно"""
    if isinstance(v, ObjectId):
        return str(v)
    if isinstance(v, datetime):
        return v.isoformat()
    if isinstance(v, list):
        return [_norm(x) for x in v]
    if isinstance(v, dict):
        return {k: _norm(x) for k, x in v.items()}
    return v


def _num(v, default=0.0):
    try:
        return float(v)
    except Exception:
        return float(default)

def _num_ru(v, default=0.0):

    if v is None:
        return float(default)
    if isinstance(v, (int, float)):
        return float(v)
    try:
        s = str(v)
        s = s.replace("\u00A0", " ").replace(" ", "")
        s = s.replace(",", ".")
        import re
        s = re.sub(r"[^0-9.\-]+", "", s)
        if not s:
            return float(default)
        return float(s)
    except Exception:
        return float(default)


def _ensure_int(v, default=0):
    try:
        return int(v)
    except Exception:
        return int(default)


def _mk_posstr(path: list[int]) -> str:
    return ".".join(str(x) for x in path if isinstance(x, int) and x > 0)


def _pick_version(doc: dict, v: int) -> Optional[dict]:
    for rec in (doc.get("versions") or []):
        if int(rec.get("v")) == int(v):
            return rec
    return None


def _set_flat_from_version(set_dict: Dict[str, Any], version_rec: dict) -> None:
    ITEM_FIELDS = (
        "pos",
        "name",
        "sku",
        "vendor",
        "unit",
        "qty",
        "price_work",
        "price_mat",
        "note",
        "rowType",
    )
    data = (version_rec or {}).get("data", {}) if version_rec else {}
    for k in ITEM_FIELDS:
        set_dict[k] = data.get(k)
    set_dict["total"] = data.get("total")


def _calc_total(d: Dict[str, Any]) -> float:
    return _num(d.get("qty"), 0) * (_num(d.get("price_work"), 0) + _num(d.get("price_mat"), 0))

def _ascii_filename(name: str) -> str:
    name = (name or "file").replace('"', "'")
    name = re.sub(r"[^A-Za-z0-9._()-]+", "_", name).strip("._- ")
    return (name[:150] or "file")

def _normalize_header_text(s: str) -> str:
    if not isinstance(s, str):
        s = str(s)
    s = s.lower()
    s = s.replace("\n", " ").replace("\r", " ").replace("\u00a0", " ")
    s = re.sub(r"[\s\-]+", "", s)
    return s

def _find_excel_col(ws, keywords):
    if not keywords:
        return None, None

    norm_keys = [_normalize_header_text(k) for k in keywords if k]
    found_col = None
    found_row = None

    for row in range(1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row=row, column=col).value
            if val in (None, ""):
                continue
            norm_val = _normalize_header_text(val)
            for nk in norm_keys:
                if nk and nk in norm_val:
                    if found_row is None or row < found_row:
                        found_row = row
                        found_col = col
                    break

    return found_col, found_row

# vlozheniya
fs_spec_section = AsyncIOMotorGridFSBucket(db, bucket_name="spec_section_attachments")
fs_ship_section = AsyncIOMotorGridFSBucket(db, bucket_name="ship_section_attachments")

# Колонки по умолчанию для v1 всех новых разделов
DEFAULT_COLUMNS = {
    "order": ["pos", "name", "sku", "vendor", "unit", "qty", "price_work", "price_mat", "total", "note", "version"],
    "hidden": [],
}


def _section_title(sec: dict) -> str:
    t = (sec.get("title") or "").strip()
    return t or "Раздел"


def _item_title(it: dict) -> str:
    pos = (it.get("posStr") or "").strip()
    if not pos:
        p = it.get("pos")
        pos = str(p) if p not in (None, "", 0) else ""
    name = (it.get("name") or "").strip()
    if pos and name:
        return f"{pos} {name}"
    return name or pos or "Позиция"


def _section_meta(sec: dict) -> dict:
    return {
        "sectionId": str(sec.get("_id")) if sec.get("_id") else None,
        "projectId": str(sec.get("projectId")) if sec.get("projectId") else None,
        "version": int(sec.get("version", sec.get("activeVersion", 1) or 1)),
        "activeVersion": int(sec.get("activeVersion", 1)),
        "deleted": bool(sec.get("deleted", False)),
    }


def _item_meta(it: dict) -> dict:
    return {
        "itemId": str(it.get("_id")) if it.get("_id") else None,
        "sectionId": str(it.get("sectionId")) if it.get("sectionId") else None,
        "projectId": str(it.get("projectId")) if it.get("projectId") else None,
        "version": int(it.get("version", it.get("activeVersion", 1) or 1)),
        "activeVersion": int(it.get("activeVersion", 1)),
        "deleted": bool(it.get("deleted", False)),
        "path": it.get("path") or [],
    }

# СЕКЦИИ

@router.get("/api/projects/{project_id}/spec/sections")
async def spec_sections_list(
    project_id: str,
    deleted: Optional[int] = Query(None),
    user=Depends(auth.get_current_user),
):
    q = {
        "tenantId": _oid(user["tenantId"]),
        "projectId": _oid(project_id),
    }
    if deleted is not None:
        q["deleted"] = bool(int(deleted))

    items = (
        await db["spec_sections"]
        .find(q)
        .sort([("order", 1), ("createdAt", 1)])
        .to_list(100000)
    )

    out = []
    for s in items:
        av = int(s.get("activeVersion", 1) or 1)
        svers = s.get("versions") or []
        rec = next((x for x in svers if int(x.get("v")) == av), None)
        out.append(
            {
                **_norm(s),
                "version": int(s.get("version", av)),
                "activeVersion": av,
                "versions": [{"v": int(x.get("v")), "savedAt": x.get("savedAt")} for x in svers],
                "columns": (rec or {}).get("columns") or DEFAULT_COLUMNS,
            }
        )
    return {"items": out}


@router.post("/api/projects/{project_id}/spec/sections")
async def spec_sections_create(
    project_id: str,
    payload: dict = Body(...),
    user=Depends(auth.get_current_user),
):
    now = datetime.utcnow()
    title = (payload.get("title") or "").strip() or "Раздел"

    last = (
        await db["spec_sections"]
        .find({"tenantId": _oid(user["tenantId"]), "projectId": _oid(project_id)})
        .sort([("order", -1)])
        .limit(1)
        .to_list(1)
    )
    next_order = int((last[0]["order"] if last else 0) or 0) + 1

    doc = {
        "tenantId": _oid(user["tenantId"]),
        "projectId": _oid(project_id),
        "title": title,
        "order": next_order,
        "comment": "",
        "deleted": False,
        "versions": [
            {
                "v": 1,
                "columns": DEFAULT_COLUMNS,
                "savedAt": now,
                "savedBy": str(user.get("_id")),
            }
        ],
        "activeVersion": 1,
        "version": 1,
        "createdAt": now,
        "updatedAt": now,
    }
    res = await db["spec_sections"].insert_one(doc)
    created = await db["spec_sections"].find_one({"_id": res.inserted_id})

    av = int(created.get("activeVersion", 1))
    cols = (_pick_version(created, av) or {}).get("columns") or DEFAULT_COLUMNS
    try:
        await log_project_action(
            db,
            user,
            project_id=project_id,
            action="spec.section.create",
            entity="spec.section",
            entity_id=str(res.inserted_id),
            message=f'Создан раздел спецификации «{title}»',
            meta={"sectionId": str(res.inserted_id)},
        )
    except Exception:
        pass

    return {
        **_norm(created),
        "columns": cols,
        "versions": [{"v": 1, "savedAt": now}],
    }

@router.patch("/api/spec/sections/{section_id}")
async def spec_sections_update(
    section_id: str,
    payload: dict = Body(...),
    user=Depends(auth.get_current_user),
):

    now = datetime.utcnow()
    sec = await db["spec_sections"].find_one({"_id": _oid(section_id), "tenantId": _oid(user["tenantId"])})
    if not sec:
        raise HTTPException(404, "Section not found")
    before_sec = dict(sec)

    # переключение версии раздела
    if "setActiveVersion" in payload:
        v = int(payload["setActiveVersion"])
        if not _pick_version(sec, v):
            raise HTTPException(404, f"Section version v{v} not found")

        await db["spec_sections"].update_one(
            {"_id": sec["_id"]},
            {"$set": {"activeVersion": v, "updatedAt": now}},
        )

        cursor = db["spec_items"].find(
            {"tenantId": _oid(user["tenantId"]), "sectionId": sec["_id"]}
        )
        async for it in cursor:
            rec = _pick_version(it, v)
            if not rec:
                continue
            upd: Dict[str, Any] = {"activeVersion": v, "updatedAt": now}
            _set_flat_from_version(upd, rec)
            await db["spec_items"].update_one({"_id": it["_id"]}, {"$set": upd})

        after = await db["spec_sections"].find_one({"_id": sec["_id"]})
        av = int(after.get("activeVersion", v))
        cols = (_pick_version(after, av) or {}).get("columns") or DEFAULT_COLUMNS

        try:
            diff = make_diff(before_sec, after)
            if diff:
                await log_project_action(
                    db,
                    user,
                    project_id=after.get("projectId"),
                    action="spec.section.set_active_version",
                    entity="spec.section",
                    entity_id=str(after["_id"]),
                    message=f'Переключена активная версия раздела «{_section_title(after)}» на v{av}',
                    diff=diff,
                    meta={**_section_meta(after), "setActiveVersion": v},
                )
        except Exception:
            pass

        return {
            **_norm(after),
            "columns": cols,
            "versions": [{"v": int(x.get("v")), "savedAt": x.get("savedAt")} for x in (after.get("versions") or [])],
        }

    # удалить конкретную версию раздела
    if "deleteVersion" in payload:
        v = int(payload["deleteVersion"])
        versions = sec.get("versions") or []
        if len(versions) <= 1:
            raise HTTPException(400, "Can't delete the only version")
        if not any(int(x.get("v")) == v for x in versions):
            raise HTTPException(404, f"Section version v{v} not found")

        # у всех позиций останется хотя бы одна версия
        cursor = db["spec_items"].find({"tenantId": _oid(user["tenantId"]), "sectionId": sec["_id"]})
        async for it in cursor:
            vers = it.get("versions") or []
            if len(vers) <= 1 and int(vers[0].get("v", 0)) == v:
                raise HTTPException(400, "Some items have only this version; can't delete")

        # сама секция
        new_versions = [x for x in versions if int(x.get("v")) != v]
        new_active = int(sec.get("activeVersion", 1))
        if new_active == v:
            new_active = max(int(x.get("v")) for x in new_versions)
        new_latest = max(int(x.get("v")) for x in new_versions)

        await db["spec_sections"].update_one(
            {"_id": sec["_id"]},
            {"$set": {"versions": new_versions, "activeVersion": new_active, "version": new_latest, "updatedAt": now}}
        )

        # позиции раздела
        cursor = db["spec_items"].find({"tenantId": _oid(user["tenantId"]), "sectionId": sec["_id"]})
        async for it in cursor:
            its = it.get("versions") or []
            filt = [x for x in its if int(x.get("v")) != v]
            if len(filt) == len(its):
                continue

            upd: Dict[str, Any] = {"versions": filt, "updatedAt": now}
            upd["version"] = max(int(x.get("v")) for x in filt)
            active_v = int(it.get("activeVersion", 1))
            if active_v == v:
                new_av = max(int(x.get("v")) for x in filt)
                upd["activeVersion"] = new_av
                rec = _pick_version({"versions": filt}, new_av)
                _set_flat_from_version(upd, rec)

            await db["spec_items"].update_one({"_id": it["_id"]}, {"$set": upd})

        after = await db["spec_sections"].find_one({"_id": sec["_id"]})
        av = int(after.get("activeVersion", new_active))
        cols = (_pick_version(after, av) or {}).get("columns") or DEFAULT_COLUMNS
        try:
            diff = make_diff(before_sec, after)
            if diff:
                await log_project_action(
                    db,
                    user,
                    project_id=after.get("projectId"),
                    action="spec.section.delete_version",
                    entity="spec.section",
                    entity_id=str(after["_id"]),
                    message=f'Удалена версия v{v} раздела «{_section_title(after)}»',
                    diff=diff,
                    meta={**_section_meta(after), "deleteVersion": v},
                )
        except Exception:
            pass
        return {
            **_norm(after),
            "columns": cols,
            "versions": [{"v": int(x.get("v")), "savedAt": x.get("savedAt")} for x in (after.get("versions") or [])],
        }

    # массовый коммит раздела (новая версия)
    if payload.get("commit"):
        new_columns = payload.get("columns") or (_pick_version(sec, int(sec.get("activeVersion", 1))) or {}).get("columns") or DEFAULT_COLUMNS
        incoming_items: List[dict] = payload.get("items") or []

        by_id: Dict[str, dict] = {}
        new_items_payload: List[dict] = []
        for it in incoming_items:
            if it.get("_id"):
                by_id[str(it["_id"])] = it
            else:
                new_items_payload.append(it)

        new_v = int(sec.get("version", 1)) + 1

        await db["spec_sections"].update_one(
            {"_id": sec["_id"]},
            {
                "$set": {"version": new_v, "activeVersion": new_v, "updatedAt": now},
                "$push": {
                    "versions": {
                        "v": new_v,
                        "columns": new_columns,
                        "savedAt": now,
                        "savedBy": str(user.get("_id")),
                    }
                },
            },
        )

        cursor = db["spec_items"].find({"tenantId": _oid(user["tenantId"]), "sectionId": sec["_id"], "deleted": {"$ne": True}})
        async for it in cursor:
            incoming = by_id.get(str(it["_id"]))
            if incoming:
                d = {
                    "pos": str(incoming.get("pos") or it.get("pos") or ""),
                    "name": incoming.get("name", it.get("name")),
                    "sku": incoming.get("sku", it.get("sku")),
                    "vendor": incoming.get("vendor", it.get("vendor")),
                    "unit": incoming.get("unit", it.get("unit")),
                    "qty": _num(incoming.get("qty", it.get("qty", 0))),
                    "price_work": _num(incoming.get("price_work", it.get("price_work", 0))),
                    "price_mat": _num(incoming.get("price_mat", it.get("price_mat", 0))),
                    "note": incoming.get("note", it.get("note")),
                    "rowType": incoming.get("rowType", it.get("rowType", "item")),
                }
            else:
                prev_v = int(sec.get("activeVersion", 1))
                prev_rec = _pick_version(it, prev_v) or {}
                base = prev_rec.get("data", {}) or {}
                d = {
                    "pos": str(base.get("pos") or it.get("pos") or ""),
                    "name": base.get("name", it.get("name")),
                    "sku": base.get("sku", it.get("sku")),
                    "vendor": base.get("vendor", it.get("vendor")),
                    "unit": base.get("unit", it.get("unit")),
                    "qty": _num(base.get("qty", it.get("qty", 0))),
                    "price_work": _num(base.get("price_work", it.get("price_work", 0))),
                    "price_mat": _num(base.get("price_mat", it.get("price_mat", 0))),
                    "note": base.get("note", it.get("note")),
                    "rowType": base.get("rowType", it.get("rowType", "item")),
                }

            d["total"] = _calc_total(d)

            updates = {
                "version": new_v,
                "activeVersion": new_v,
                "updatedAt": now,
            }
            for k in ["pos", "name", "sku", "vendor", "unit", "qty", "price_work", "price_mat", "note", "rowType"]:
                updates[k] = d.get(k)
            updates["total"] = d["total"]

            await db["spec_items"].update_one(
                {"_id": it["_id"]},
                {
                    "$set": updates,
                    "$push": {"versions": {"v": new_v, "data": d, "savedAt": now, "savedBy": str(user.get("_id"))}},
                },
            )

        for payload_item in new_items_payload:
            await _create_item_core(user, sec, payload_item, new_v, now)

        after = await db["spec_sections"].find_one({"_id": sec["_id"]})
        av = int(after.get("activeVersion", new_v))
        cols = (_pick_version(after, av) or {}).get("columns") or DEFAULT_COLUMNS
        try:
            diff = make_diff(before_sec, after)
            if diff:
                await log_project_action(
                    db,
                    user,
                    project_id=after.get("projectId"),
                    action="spec.section.commit",
                    entity="spec.section",
                    entity_id=str(after["_id"]),
                    message=f'Сохранена новая версия v{new_v} раздела «{_section_title(after)}»',
                    diff=diff,
                    meta={**_section_meta(after), "newVersion": new_v},
                )
        except Exception:
            pass
        return {
            **_norm(after),
            "columns": cols,
            "versions": [{"v": int(x.get("v")), "savedAt": x.get("savedAt")} for x in (after.get("versions") or [])],
        }
    
    # обновление колонок активной версии без создания v+1
    if "columns" in payload:
        new_columns = payload.get("columns") or DEFAULT_COLUMNS
        versions = sec.get("versions") or []
        av = int(sec.get("activeVersion", sec.get("version", 1) or 1))

        found = False
        for rec in versions:
            if int(rec.get("v")) == av:
                rec["columns"] = new_columns
                rec["savedAt"] = now
                rec["savedBy"] = str(user.get("_id"))
                found = True
                break
        if not found:
            versions.append(
                {
                    "v": av,
                    "columns": new_columns,
                    "savedAt": now,
                    "savedBy": str(user.get("_id")),
                }
            )

        await db["spec_sections"].update_one(
            {"_id": sec["_id"]},
            {"$set": {"versions": versions, "updatedAt": now}},
        )

        after = await db["spec_sections"].find_one({"_id": sec["_id"]})
        av2 = int(after.get("activeVersion", av))
        cols = (_pick_version(after, av2) or {}).get("columns") or DEFAULT_COLUMNS

        try:
            diff = make_diff(before_sec, after)
            if diff:
                await log_project_action(
                    db,
                    user,
                    project_id=after.get("projectId"),
                    action="spec.section.update_columns",
                    entity="spec.section",
                    entity_id=str(after["_id"]),
                    message=f'Обновлены столбцы раздела «{_section_title(after)}» для v{av2}',
                    diff=diff,
                    meta=_section_meta(after),
                )
        except Exception:
            pass

        return {
            **_norm(after),
            "columns": cols,
            "versions": [{"v": int(x.get("v")), "savedAt": x.get("savedAt")} for x in (after.get("versions") or [])],
        }

    # обычные поля + каскад архива
    updates: Dict[str, Any] = {}
    cascade_deleted: Optional[bool] = None
    if "title" in payload:
        updates["title"] = (payload.get("title") or "").strip()
    if "order" in payload:
        try:
            updates["order"] = int(payload.get("order") or 0)
        except Exception:
            pass
    if "deleted" in payload:
        updates["deleted"] = bool(payload["deleted"])
        cascade_deleted = updates["deleted"]
    if "comment" in payload:
        updates["comment"] = payload.get("comment") or ""

    if not updates:
        return _norm(sec)

    updates["updatedAt"] = datetime.utcnow()
    await db["spec_sections"].update_one({"_id": sec["_id"]}, {"$set": updates})

    if cascade_deleted is not None:
        now2 = datetime.utcnow()
        await db["spec_items"].update_many(
            {"tenantId": _oid(user["tenantId"]), "sectionId": sec["_id"]},
            {"$set": {"deleted": cascade_deleted, "updatedAt": now2}},
        )

    after = await db["spec_sections"].find_one({"_id": sec["_id"]})
    av = int(after.get("activeVersion", 1))
    cols = (_pick_version(after, av) or {}).get("columns") or DEFAULT_COLUMNS
    try:
        diff = make_diff(before_sec, after)
        if diff:
            await log_project_action(
                db,
                user,
                project_id=after.get("projectId"),
                action="spec.section.update",
                entity="spec.section",
                entity_id=str(after["_id"]),
                message=f'Обновлён раздел спецификации «{_section_title(after)}»',
                diff=diff,
                meta=_section_meta(after),
            )
    except Exception:
        pass
    return {**_norm(after), "columns": cols}


@router.delete("/api/spec/sections/{section_id}")
async def spec_sections_delete_forever(
    section_id: str,
    user=Depends(auth.get_current_user),
):
    sec = await db["spec_sections"].find_one({"_id": _oid(section_id), "tenantId": _oid(user["tenantId"])})
    if not sec:
        raise HTTPException(404, "Section not found")

    res_items = await db["spec_items"].delete_many({"tenantId": _oid(user["tenantId"]), "sectionId": sec["_id"]})
    res_sec = await db["spec_sections"].delete_one({"_id": sec["_id"]})
    try:
        await log_project_action(
            db,
            user,
            project_id=sec.get("projectId"),
            action="spec.section.delete",
            entity="spec.section",
            entity_id=str(section_id),
            message=f'Полностью удалён раздел спецификации «{_section_title(sec)}»',
            meta={
                **_section_meta(sec),
                "itemsDeleted": res_items.deleted_count,
            },
        )
    except Exception:
        pass

    return {"ok": True}


@router.get("/api/spec/sections/{section_id}/versions")
async def spec_section_versions(section_id: str, user=Depends(auth.get_current_user)):
    sec = await db["spec_sections"].find_one({"_id": _oid(section_id), "tenantId": _oid(user["tenantId"])})
    if not sec:
        raise HTTPException(404, "Section not found")
    return {
        "_id": str(sec["_id"]),
        "activeVersion": int(sec.get("activeVersion", 1)),
        "version": int(sec.get("version", 1)),
        "versions": _norm(sec.get("versions") or []),
    }

@router.get("/api/spec/sections/{section_id}/export")
async def spec_section_export_excel(
    section_id: str,
    user=Depends(auth.get_current_user),
):
    """
    Экспорт раздела спецификации в Excel.

    Один лист "Раздел":
      заголовки групп (материалы, кабели и т.п.)
      позиции
      сразу под позицией — строки работ

      Колонки:
        №, Наименование, Артикул, Поставщик, Единица измерения, Кол-во, Цена работы, Цена материалов, Стоимость, Примечание
    """
    # находим раздел
    sec = await db["spec_sections"].find_one(
        {"_id": _oid(section_id), "tenantId": _oid(user["tenantId"])}
    )
    if not sec:
        raise HTTPException(404, "Section not found")

    # версия, по которой делаем срез
    v = int(sec.get("activeVersion", sec.get("version", 1) or 1))

    tenant_id = _oid(user["tenantId"])
    section_oid = sec["_id"]
    project_oid = sec.get("projectId")

    # берем все не удалённые позиции раздела
    items = (
        await db["spec_items"]
        .find(
            {
                "tenantId": tenant_id,
                "sectionId": section_oid,
                "deleted": {"$ne": True},
            }
        )
        .to_list(100000)
    )

    rows: List[dict] = []

    for it in items:
        rec = _pick_version(it, v)
        if not rec:
            continue

        data = rec.get("data") or {}
        path = it.get("path") or []
        row_type = data.get("rowType") or it.get("rowType") or "item"
        total_val = _calc_total(data) if row_type != "header" else 0.0

        rows.append(
            {
                "itemId": str(it["_id"]),
                "path": path,
                "rowType": row_type,
                "name": data.get("name") or "",
                "sku": data.get("sku") or "",
                "vendor": data.get("vendor") or "",
                "unit": data.get("unit") or "",
                "qty": data.get("qty") or 0,
                "price_work": data.get("price_work") or 0,
                "price_mat": data.get("price_mat") or 0,
                "total": total_val,
                "note": data.get("note") or "",
            }
        )

    # сортировка
    rows.sort(key=lambda r: r["path"])

    # нумерация внутри заголовков
    current_group_key = None
    current_index = 0
    for r in rows:
        path = r.get("path") or []
        if r["rowType"] == "header":
            current_group_key = tuple(path)
            current_index = 0
            r["posDisplay"] = ""
        else:
            parent_key = tuple(path[:-1]) if len(path) > 1 else None
            if parent_key and parent_key == current_group_key:
                current_index += 1
                r["posDisplay"] = str(current_index)
            else:
                r["posDisplay"] = _mk_posstr(path)

    # работы по позициям
    works = (
        await db["spec_item_works"]
        .find(
            {
                "tenantId": tenant_id,
                "projectId": project_oid,
                "sectionId": section_oid,
                "deleted": {"$ne": True},
            }
        )
        .sort([("itemId", 1), ("order", 1), ("createdAt", 1)])
        .to_list(100000)
    )

    from collections import defaultdict

    works_by_item: Dict[str, List[dict]] = defaultdict(list)
    for w in works:
        works_by_item[str(w.get("itemId"))].append(w)
    
    # формируем Excel
    try:
        wb = Workbook()
        ws_spec = wb.active
        ws_spec.title = "Раздел"

        spec_headers = [
            "№",
            "Наименование",
            "Артикул",
            "Поставщик",
            "Единица измерения",
            "Кол-во",
            "Цена работы",
            "Цена материалов",
            "Стоимость",
            "Примечание",
        ]
        ws_spec.append(spec_headers)

        header_rows_excel: List[int] = []
        work_rows_excel: List[int] = []

        for r in rows:
            if r["rowType"] == "header":
                ws_spec.append(
                    [
                        "",        # №
                        r["name"], # Наименование (заголовок)
                        "",
                        "",
                        "",
                        "",
                        "",
                        "",
                        "",
                        "",
                    ]
                )
                header_rows_excel.append(ws_spec.max_row)
            else:
                # позиция
                ws_spec.append(
                    [
                        r.get("posDisplay", ""),
                        r["name"],
                        r["sku"],
                        r["vendor"],
                        r["unit"],
                        r["qty"],
                        r["price_work"],
                        r["price_mat"],
                        r["total"],
                        r["note"],
                    ]
                )

                # работы под позицией
                for w in works_by_item.get(r["itemId"], []):
                    work_name = (w.get("name") or "").strip()
                    qty_plan = w.get("qty_plan") or 0

                    ws_spec.append(
                        [
                            "",
                            work_name,
                            "",
                            "",
                            "",
                            qty_plan,
                            "",
                            "",
                            "",
                            "",
                        ]
                    )
                    work_rows_excel.append(ws_spec.max_row)

        # оформление
        from openpyxl.styles import Border, Side, Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter

        thin = Side(style="thin", color="000000")
        border_all = Border(left=thin, right=thin, top=thin, bottom=thin)

        spec_max_row = ws_spec.max_row
        spec_max_col = ws_spec.max_column

        widths = {1: 5, 2: 45, 3: 15, 4: 18, 5: 16, 6: 10, 7: 14, 8: 14, 9: 16, 10: 25}
        for col_idx, width in widths.items():
            ws_spec.column_dimensions[get_column_letter(col_idx)].width = width

        for row in ws_spec.iter_rows(
            min_row=1, max_row=spec_max_row, min_col=1, max_col=spec_max_col
        ):
            for cell in row:
                cell.border = border_all
                cell.alignment = Alignment(
                    wrap_text=True,
                    vertical="top",
                    horizontal=cell.alignment.horizontal or "left",
                )

        # шапка
        for cell in ws_spec[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )

        # заголовки групп
        for row_idx in header_rows_excel:
            ws_spec.merge_cells(start_row=row_idx, start_column=2, end_row=row_idx, end_column=5)
            hdr_cell = ws_spec.cell(row=row_idx, column=2)
            hdr_cell.font = Font(bold=True)
            hdr_cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )

        # подсветка строк работ
        work_fill_blue = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
        for row_idx in work_rows_excel:
            for col in range(2, spec_max_col + 1):
                ws_spec.cell(row=row_idx, column=col).fill = work_fill_blue

        # отдаём файл
        stream = BytesIO()
        wb.save(stream)
        stream.seek(0)
    except Exception as e:
        raise HTTPException(500, f"Ошибка при формировании Excel файла: {e}")

    import re
    safe_title = _section_title(sec).replace('"', "'").replace("\n", " ")
    ascii_title = re.sub(r"[^A-Za-z0-9_.-]+", "_", safe_title).strip("_") or "section"
    filename = f"{ascii_title}_v{v}.xlsx"

    headers_resp = {"Content-Disposition": f'attachment; filename="{filename}"'}

    try:
        await log_project_action(
            db,
            user,
            project_id=sec.get("projectId"),
            action="spec.section.export_excel",
            entity="spec.section",
            entity_id=str(sec["_id"]),
            message=f'Экспортирована спецификация раздела «{_section_title(sec)}» в Excel',
            meta=_section_meta(sec),
        )
    except Exception:
        pass

    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers_resp,
    )

@router.post("/api/spec/sections/{section_id}/import")
async def spec_section_import_excel(
    section_id: str,
    file: UploadFile = File(...),
    user=Depends(auth.get_current_user),
):
    sec = await db["spec_sections"].find_one(
        {"_id": _oid(section_id), "tenantId": _oid(user["tenantId"])}
    )
    if not sec:
        raise HTTPException(404, "Section not found")

    filename = file.filename or ""
    if not filename.lower().endswith(".xlsx"):
        raise HTTPException(400, "Ожидается .xlsx файл")

    try:
        content = await file.read()
        if not content:
            raise HTTPException(400, "Пустой файл")

        stream = BytesIO(content)
        wb = load_workbook(stream, data_only=True)
        ws = wb.active
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(400, f"Не удалось прочитать Excel: {e}")

    if ws.max_row < 2:
        raise HTTPException(400, "В файле нет данных")

    # читаем заголовки
    header_row = 1
    header_raw: Dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=header_row, column=col).value
        if val is None:
            continue
        name = str(val).strip()
        if not name:
            continue
        header_raw[name] = col

    if not header_raw:
        raise HTTPException(400, "Не найдена строка заголовков в первой строке")

    def norm_header(h: str) -> str:
        return h.strip().lower().replace("ё", "е")

    synonyms = {
        "posStr": [
            "№",
            "no",
            "номер",
            "n",
            "#",
            "№п/п",
            "№ п/п",
            "номер п/п",
        ],
        "name": [
            "наименование",
            "наименования видов работ",
            "наименование видов работ",
            "name",
        ],
        "sku": [
            "артикул",
            "sku",
        ],
        "vendor": [
            "поставщик",
            "vendor",
        ],
        "unit": [
            "единица",
            "единица измерения",
            "ед.изм.",
            "ед. изм.",
            "ед изм",
            "unit",
        ],
        "qty": [
            "кол-во",
            "количество",
            "qty",
        ],
        "price_work": [
            "цена работы",
            "цена работ",
            "в том числе работ",
            "работ",
        ],
        "price_mat": [
            "цена материалов",
            "цена материлов",
            "в том числе материалов",
            "материалов",
        ],
        "note": [
            "примечание",
            "комментарий",
        ],
        "rowType": [
            "тип строки",
            "type",
        ],
    }

    col_idx: Dict[str, int] = {}
    for raw, col in header_raw.items():
        nh = norm_header(raw)
        for internal, names in synonyms.items():
            if any(nh == norm_header(x) for x in names):
                if internal not in col_idx:
                    col_idx[internal] = col
                break
    
    if "price_work" not in col_idx and "price_mat" not in col_idx:
        for raw, col in header_raw.items():
            nh = norm_header(raw)
            if "цена единиц" in nh or nh.startswith("цена, руб") or nh.startswith("цена "):
                col_idx["price_work"] = col
                break

    if "posStr" not in col_idx:
        raise HTTPException(
            400, "Не найдена колонка с номером позиции («№»)"
        )
    if "name" not in col_idx:
        raise HTTPException(
            400, "Не найдена колонка «Наименование»"
        )

    # парсим строки
    parsed_rows: List[dict] = []

    for row in range(2, ws.max_row + 1):
        is_empty = True
        for col in range(1, ws.max_column + 1):
            if ws.cell(row=row, column=col).value not in (None, ""):
                is_empty = False
                break
        if is_empty:
            continue

        def get_cell(internal: str):
            cidx = col_idx.get(internal)
            if not cidx:
                return None
            return ws.cell(row=row, column=cidx).value

        raw_pos = get_cell("posStr")
        raw_name = get_cell("name")

        pos_str = str(raw_pos).strip() if raw_pos is not None else ""
        name = str(raw_name).strip() if raw_name is not None else ""

        # Если нет и номера, и наименования - пропускаем
        if not pos_str and not name:
            continue

        if not pos_str:
            raise HTTPException(
                400,
                f"Строка {row}: в колонке «№» должен быть номер вида '1' или '1.2'",
            )

        # парсим иерархический номер
        try:
            parts = [
                int(p)
                for p in str(pos_str).replace(",", ".").split(".")
                if p.strip()
            ]
        except Exception:
            raise HTTPException(
                400,
                f"Строка {row}: некорректный номер «{pos_str}»",
            )

        if not parts:
            raise HTTPException(
                400,
                f"Строка {row}: некорректный номер «{pos_str}»",
            )

        parent_key = ".".join(str(x) for x in parts[:-1]) if len(parts) > 1 else None
        this_key = ".".join(str(x) for x in parts)

        raw_type = get_cell("rowType")
        rt = str(raw_type).strip().lower() if raw_type is not None else ""
        if rt in ("header", "заголовок"):
            row_type = "header"
        else:
            row_type = "item"

        qty_raw = get_cell("qty")
        price_work_raw = get_cell("price_work")
        price_mat_raw = get_cell("price_mat")

        # собираем данные позиции
        data = {
            "name": name,
            "sku": (get_cell("sku") or "") if "sku" in col_idx else "",
            "vendor": (get_cell("vendor") or "") if "vendor" in col_idx else "",
            "unit": (get_cell("unit") or "") if "unit" in col_idx else "",
            "qty": 0 if row_type == "header" else _num(qty_raw, 0),
            "price_work": 0 if row_type == "header" else _num(price_work_raw, 0),
            "price_mat": 0 if row_type == "header" else _num(price_mat_raw, 0),
            "note": (get_cell("note") or "") if "note" in col_idx else "",
            "rowType": row_type,
        }

        parsed_rows.append(
            {
                "row_number": row,
                "parts": parts,
                "parent_key": parent_key,
                "this_key": this_key,
                "data": data,
            }
        )

    if not parsed_rows:
        raise HTTPException(400, "В файле не найдено ни одной позиции")

    # сортируем так, чтобы родители шли раньше детей
    parsed_rows.sort(key=lambda r: (len(r["parts"]), r["parts"]))

    now = datetime.utcnow()
    v_for_snapshot = int(sec.get("activeVersion", sec.get("version", 1) or 1))

    # помечаем старые позиции раздела как deleted
    await db["spec_items"].update_many(
        {
            "tenantId": _oid(user["tenantId"]),
            "sectionId": sec["_id"],
            "deleted": {"$ne": True},
        },
        {"$set": {"deleted": True, "updatedAt": now}},
    )

    created_map: Dict[str, ObjectId] = {}

    # создаём новые позиции согласно иерархии
    for row in parsed_rows:
        parent_id = None
        if row["parent_key"]:
            parent_id = created_map.get(row["parent_key"])
            if not parent_id:
                raise HTTPException(
                    400,
                    f"Строка {row['row_number']}: не найден родитель для номера «{row['parent_key']}»",
                )

        payload_item = dict(row["data"])
        payload_item["parentId"] = str(parent_id) if parent_id else None
        payload_item["pos"] = row["parts"][-1]

        doc = await _create_item_core(
            user, sec, payload_item, v_for_snapshot, now
        )
        created_map[row["this_key"]] = doc["_id"]

    after = await db["spec_sections"].find_one({"_id": sec["_id"]})
    av = int(after.get("activeVersion", v_for_snapshot))
    cols = (_pick_version(after, av) or {}).get("columns") or DEFAULT_COLUMNS

    try:
        await log_project_action(
            db,
            user,
            project_id=after.get("projectId"),
            action="spec.section.import_excel",
            entity="spec.section",
            entity_id=str(after["_id"]),
            message=f'Импортирована спецификация раздела «{_section_title(after)}» из Excel',
            meta=_section_meta(after),
        )
    except Exception:
        pass

    return {
        **_norm(after),
        "columns": cols,
        "versions": [
            {"v": int(x.get("v")), "savedAt": x.get("savedAt")}
            for x in (after.get("versions") or [])
        ],
    }

# ПОЗИЦИИ

ITEM_FIELDS = (
    "pos",
    "name",
    "sku",
    "vendor",
    "unit",
    "qty",
    "price_work",
    "price_mat",
    "note",
)


def _active_payload(item: dict) -> dict:
    av = int(item.get("activeVersion", 1))
    snap = next((x for x in (item.get("versions") or []) if int(x.get("v")) == av), None) or {}
    data = snap.get("data", {})
    flat = {k: data.get(k) for k in ITEM_FIELDS}
    flat["total"] = data.get("total")
    flat["version"] = int(item.get("version", av))
    flat["activeVersion"] = av
    flat["versions"] = [{"v": int(x.get("v")), "savedAt": x.get("savedAt")} for x in (item.get("versions") or [])]
    return flat

@router.get("/api/projects/{project_id}/spec/items")
async def spec_items_list(
    project_id: str,
    deleted: Optional[int] = Query(None),
    sectionId: Optional[str] = Query(None),
    user=Depends(auth.get_current_user),
):
    q = {"tenantId": _oid(user["tenantId"]), "projectId": _oid(project_id)}
    if deleted is not None:
        q["deleted"] = bool(int(deleted))
    if sectionId:
        q["sectionId"] = _oid(sectionId)

    items = (
        await db["spec_items"]
        .find(q)
        .sort([("sectionOrder", 1), ("createdAt", 1)])
        .to_list(100000)
    )
    out: List[dict] = []
    for it in items:
        flat = _active_payload(it)
        out.append({**_norm(it), **flat, "versions": flat["versions"]})
    return {"items": out}

async def _create_item_core(user, sec: dict, payload: dict, version_for_snapshot: int, now: datetime):
    section_id = sec["_id"]
    project_id = sec["projectId"]

    parent_id = payload.get("parentId")
    parent = None
    if parent_id:
        parent = await db["spec_items"].find_one(
            {
                "_id": _oid(parent_id),
                "tenantId": _oid(user["tenantId"]),
                "sectionId": _oid(section_id),
                "deleted": {"$ne": True},
            }
        )
        if not parent:
            raise HTTPException(404, "Parent item not found")

    level = (parent.get("level", 0) + 1) if parent else 1
    if level > 3:
        raise HTTPException(400, "Max depth is 3")

    siblings_q = {
        "tenantId": _oid(user["tenantId"]),
        "projectId": _oid(project_id),
        "sectionId": _oid(section_id),
        "parentId": parent["_id"] if parent else None,
        "deleted": {"$ne": True},
    }

    siblings_count = await db["spec_items"].count_documents(siblings_q)

    requested_pos = _ensure_int(payload.get("pos"), 0)
    next_pos = requested_pos if requested_pos > 0 else (siblings_count + 1)

    parent_path = parent.get("path", []) if parent else []
    path = list(parent_path) + [next_pos]
    posStr = _mk_posstr(path)

    d = {k: payload.get(k) for k in ITEM_FIELDS}
    d["rowType"] = payload.get("rowType") or "item"
    d["qty"] = _num(d.get("qty"), 1)
    d["price_work"] = _num(d.get("price_work"), 0)
    d["price_mat"] = _num(d.get("price_mat"), 0)
    d["total"] = _calc_total(d)

    doc = {
        "tenantId": _oid(user["tenantId"]),
        "projectId": project_id,
        "sectionId": section_id,
        "sectionOrder": int(sec.get("order") or 0),
        "parentId": parent["_id"] if parent else None,
        "level": level,
        "pos": next_pos,
        "path": path,
        "posStr": posStr,
        "deleted": False,
        "versions": [
            {
                "v": int(version_for_snapshot or 1),
                "data": d,
                "savedAt": now,
                "savedBy": str(user.get("_id")),
            }
        ],
        "activeVersion": int(version_for_snapshot or 1),
        "version": int(version_for_snapshot or 1),
        **d,
        "createdAt": now,
        "updatedAt": now,
    }
    await db["spec_items"].insert_one(doc)
    return doc


@router.post("/api/projects/{project_id}/spec/items")
async def spec_items_create(
    project_id: str,
    payload: dict = Body(...),
    user=Depends(auth.get_current_user),
):
    now = datetime.utcnow()
    section_id = payload.get("sectionId")
    if not section_id:
        raise HTTPException(400, "sectionId is required")

    sec = await db["spec_sections"].find_one({"_id": _oid(section_id), "tenantId": _oid(user["tenantId"])})
    if not sec:
        raise HTTPException(404, "Section not found")

    v_for_snapshot = int(sec.get("activeVersion", 1))
    doc = await _create_item_core(user, sec, payload, v_for_snapshot, now)
    created = await db["spec_items"].find_one({"_id": doc.get("_id")})
    flat = _active_payload(created)
    try:
        await log_project_action(
            db,
            user,
            project_id=created.get("projectId"),
            action="spec.item.create",
            entity="spec.item",
            entity_id=str(created["_id"]),
            message=f'Создана позиция спецификации «{_item_title(created)}»',
            meta=_item_meta(created),
        )
    except Exception:
        pass
    return {**_norm(created), **flat, "versions": flat["versions"]}


@router.patch("/api/spec/items/{item_id}")
async def spec_items_update(
    item_id: str,
    payload: dict = Body(...),
    user=Depends(auth.get_current_user),
):
    now = datetime.utcnow()
    it = await db["spec_items"].find_one({"_id": _oid(item_id), "tenantId": _oid(user["tenantId"])})
    if not it:
        raise HTTPException(404, "Item not found")
    before = dict(it)

    updates: Dict[str, Any] = {}

    if "deleted" in payload:
        flag = bool(payload["deleted"])
        updates["deleted"] = flag
        if payload.get("cascade"):
            it_path = it.get("path", [])
            await db["spec_items"].update_many(
                {
                    "tenantId": _oid(user["tenantId"]),
                    "sectionId": it["sectionId"],
                    "path": {"$gte": it_path, "$lte": it_path + [10**9]},
                },
                {"$set": {"deleted": flag, "updatedAt": now}},
            )

    if "sectionId" in payload and payload.get("sectionId"):
        sec = await db["spec_sections"].find_one({"_id": _oid(payload["sectionId"]), "tenantId": _oid(user["tenantId"])})
        if not sec:
            raise HTTPException(404, "Section not found")
        updates["sectionId"] = sec["_id"]
        updates["sectionOrder"] = int(sec.get("order") or 0)

    if "setActiveVersion" in payload:
        v = int(payload["setActiveVersion"])
        rec = _pick_version(it, v)
        if not rec:
            raise HTTPException(404, f"Version v{v} not found")
        updates["activeVersion"] = v
        _set_flat_from_version(updates, rec)

    if "deleteVersion" in payload:
        v = int(payload["deleteVersion"])
        versions = it.get("versions") or []
        if len(versions) <= 1:
            raise HTTPException(400, "Can't delete the only version")
        new_versions = [x for x in versions if int(x.get("v")) != v]
        if len(new_versions) == len(versions):
            raise HTTPException(404, f"Version v{v} not found")
        updates["versions"] = new_versions
        active_v = int(it.get("activeVersion", 1))
        if v == active_v:
            new_active = max(int(x.get("v")) for x in new_versions)
            updates["activeVersion"] = new_active
            rec = _pick_version({"versions": new_versions}, new_active)
            _set_flat_from_version(updates, rec)
        updates["version"] = max(int(x.get("v")) for x in new_versions)

    # обновление активной версии без новой v
    if payload.get("updateActive"):
        incoming = payload.get("data") or {}
        # Берём новые значения, а если поле не передано оставляем старое
        d = {k: incoming.get(k, it.get(k)) for k in ITEM_FIELDS}
        d["qty"] = _num(d.get("qty"), it.get("qty", 1))
        d["price_work"] = _num(d.get("price_work"), it.get("price_work", 0))
        d["price_mat"] = _num(d.get("price_mat"), it.get("price_mat", 0))
        d["total"] = _calc_total(d)

        active_v = int(it.get("activeVersion", it.get("version", 1)))
        vers = it.get("versions") or []
        for rec in vers:
            if int(rec.get("v")) == active_v:
                rec["data"] = d
                rec["savedAt"] = now
                rec["savedBy"] = str(user.get("_id"))
                break

        for k in ITEM_FIELDS:
            updates[k] = d.get(k)
        updates["total"] = d["total"]
        updates["versions"] = vers

    if payload.get("commit"):
        incoming = payload.get("data") or {}
        d = {k: incoming.get(k) for k in ITEM_FIELDS}
        d["rowType"] = incoming.get("rowType", it.get("rowType", "item"))
        d["qty"] = _num(d.get("qty"), 1)
        d["price_work"] = _num(d.get("price_work"), 0)
        d["price_mat"] = _num(d.get("price_mat"), 0)
        d["total"] = _calc_total(d)

        overwrite = payload.get("overwriteVersion")
        if overwrite is not None:
            v = int(overwrite)
            rec = _pick_version(it, v)
            if not rec:
                raise HTTPException(404, f"Version v{v} not found to overwrite")
            new_versions = []
            for x in (it.get("versions") or []):
                if int(x.get("v")) == v:
                    new_versions.append({**x, "data": d, "savedAt": now, "savedBy": str(user.get("_id"))})
                else:
                    new_versions.append(x)
            updates["versions"] = new_versions
            updates["updatedAt"] = now
            if int(it.get("activeVersion", 1)) == v:
                for k in ITEM_FIELDS:
                    updates[k] = d.get(k)
                updates["total"] = d["total"]
        else:
            new_v = int(it.get("version", 1)) + 1
            updates["version"] = new_v
            updates["activeVersion"] = new_v
            updates["versions"] = (it.get("versions") or []) + [
                {"v": new_v, "data": d, "savedAt": now, "savedBy": str(user.get("_id"))}
            ]
            for k in ITEM_FIELDS:
                updates[k] = d.get(k)
            updates["total"] = d["total"]

    if not updates:
        return {**_norm(it), **_active_payload(it)}

    updates["updatedAt"] = now
    await db["spec_items"].update_one({"_id": it["_id"]}, {"$set": updates})
    after = await db["spec_items"].find_one({"_id": it["_id"]})
    flat = _active_payload(after)
    try:
        diff = make_diff(before, after)
        if diff:
            await log_project_action(
                db,
                user,
                project_id=after.get("projectId"),
                action="spec.item.update",
                entity="spec.item",
                entity_id=str(after["_id"]),
                message=f'Обновлена позиция спецификации «{_item_title(after)}»',
                diff=diff,
                meta=_item_meta(after),
            )
    except Exception:
        pass
    return {**_norm(after), **flat, "versions": flat["versions"]}

@router.delete("/api/spec/items/{item_id}")
async def spec_items_delete_forever(
    item_id: str,
    user=Depends(auth.get_current_user),
):
    it = await db["spec_items"].find_one({"_id": _oid(item_id), "tenantId": _oid(user["tenantId"])})
    if not it:
        raise HTTPException(404, "Item not found")
    await db["spec_items"].delete_one({"_id": it["_id"]})

    try:
        await log_project_action(
            db,
            user,
            project_id=it.get("projectId"),
            action="spec.item.delete",
            entity="spec.item",
            entity_id=str(item_id),
            message=f'Полностью удалена позиция спецификации «{_item_title(it)}»',
            meta=_item_meta(it),
        )
    except Exception:
        pass
    return {"ok": True}


@router.get("/api/spec/items/{item_id}")
async def spec_item_get(
    item_id: str,
    user=Depends(auth.get_current_user),
):
    it = await db["spec_items"].find_one({"_id": _oid(item_id), "tenantId": _oid(user["tenantId"])})
    if not it:
        raise HTTPException(404, "Item not found")
    flat = _active_payload(it)
    vers_full = [
        {"v": int(x.get("v")), "savedAt": x.get("savedAt"), "data": x.get("data", {})}
        for x in (it.get("versions") or [])
    ]
    return {**_norm(it), **flat, "versionsFull": vers_full}

# Перерасчёт при переносе узла
async def _recalc_subtree(dbh, tenant_id, section_id, root_id, new_parent_id, new_index):
    from datetime import datetime
    now = datetime.utcnow()

    t = _oid(tenant_id)
    s = _oid(section_id)
    rid = _oid(root_id)
    pid = _oid(new_parent_id) if new_parent_id else None

    parent = None
    if pid:
        parent = await dbh.find_one({"_id": pid, "tenantId": t, "sectionId": s})
        if not parent:
            raise HTTPException(404, "Target parent not found")

    root = await dbh.find_one({"_id": rid, "tenantId": t, "sectionId": s})
    if not root:
        raise HTTPException(404, "Item not found")

    def sibs_query(parent_id):
        return {"tenantId": t, "sectionId": s, "parentId": parent_id, "deleted": {"$ne": True}}

    async def _set_pos_and_path(doc_id, new_path, new_level, new_pos, active_v):
        await dbh.update_one(
            {"_id": doc_id},
            {"$set": {
                "pos": new_pos,
                "level": new_level,
                "path": new_path,
                "posStr": _mk_posstr(new_path),
                "updatedAt": now,
            }}
        )
        await dbh.update_one(
            {"_id": doc_id},
            {"$set": {"versions.$[ver].data.pos": str(new_pos), "versions.$[ver].savedAt": now}},
            array_filters=[{"ver.v": int(active_v)}]
        )

    async def recalc_children(parent_doc: dict):
        children = await dbh.find(
            {"tenantId": t, "sectionId": s, "parentId": parent_doc["_id"], "deleted": {"$ne": True}}
        ).sort([("pos", 1), ("createdAt", 1)]).to_list(100000)

        for i, ch in enumerate(children, start=1):
            ch_new_path = parent_doc["path"] + [i]
            ch_level = parent_doc["level"] + 1
            if ch_level > 3:
                raise HTTPException(400, "Max depth is 3")
            await _set_pos_and_path(ch["_id"], ch_new_path, ch_level, i, ch.get("activeVersion", ch.get("version", 1)))
            await recalc_children({**ch, "path": ch_new_path, "level": ch_level})

    target_parent_id = parent["_id"] if parent else None
    target_sibs = await dbh.find(sibs_query(target_parent_id)).sort([("pos", 1), ("createdAt", 1)]).to_list(100000)
    target_sibs_wo_root = [x for x in target_sibs if x["_id"] != rid]

    new_index = int(new_index or 1)
    new_index = max(1, min(new_index, len(target_sibs_wo_root) + 1))

    same_parent = (root.get("parentId") == target_parent_id)

    if not same_parent:
        old_parent_id = root.get("parentId")
        old_sibs = await dbh.find(sibs_query(old_parent_id)).sort([("pos", 1), ("createdAt", 1)]).to_list(100000)
        old_sibs_wo_root = [x for x in old_sibs if x["_id"] != rid]
        for i, x in enumerate(old_sibs_wo_root, start=1):
            x_new_path = (x.get("path", [])[:-1] + [i])
            await _set_pos_and_path(x["_id"], x_new_path, x.get("level", 1), i, x.get("activeVersion", x.get("version", 1)))

    new_order = target_sibs_wo_root[:]
    new_order.insert(new_index - 1, {"_id": rid})

    base_parent_path = parent.get("path", []) if parent else []
    base_level = (parent.get("level", 0) + 1) if parent else 1
    if base_level > 3:
        raise HTTPException(400, "Max depth is 3")

    for i, doc in enumerate(new_order, start=1):
        if doc.get("_id") == rid:
            new_root_path = base_parent_path + [i]
            await dbh.update_one(
                {"_id": rid},
                {"$set": {"parentId": target_parent_id}}
            )
            await _set_pos_and_path(rid, new_root_path, base_level, i, root.get("activeVersion", root.get("version", 1)))
            await recalc_children({"_id": rid, "path": new_root_path, "level": base_level})
        else:
            cur = next((x for x in target_sibs_wo_root if x["_id"] == doc["_id"]), None)
            if not cur:
                continue
            newp = base_parent_path + [i]
            await _set_pos_and_path(cur["_id"], newp, base_level, i, cur.get("activeVersion", cur.get("version", 1)))

    return True


async def _sync_version_positions_for_parent(tenant_id, section_id, parent_id, for_version: int, saved_by: str):
    t = _oid(tenant_id)
    s = _oid(section_id)
    pid = _oid(parent_id) if parent_id else None
    now = datetime.utcnow()

    children = await db["spec_items"].find(
        {"tenantId": t, "sectionId": s, "parentId": pid, "deleted": {"$ne": True}}
    ).sort([("pos", 1), ("createdAt", 1)]).to_list(100000)

    for i, it in enumerate(children, start=1):
        vers = it.get("versions") or []
        rec = next((x for x in vers if int(x.get("v")) == int(for_version)), None)

        if rec is None:
            data = {k: it.get(k) for k in ITEM_FIELDS}
            data["pos"] = i
            data["total"] = _calc_total(data)
            vers.append({"v": int(for_version), "data": data, "savedAt": now, "savedBy": saved_by})
        else:
            data = rec.get("data") or {}
            data = {**data, "pos": i}
            data["total"] = _calc_total(data)
            rec["data"] = data
            rec["savedAt"] = now
            rec["savedBy"] = saved_by

        await db["spec_items"].update_one(
            {"_id": it["_id"]},
            {"$set": {"versions": vers, "updatedAt": now}}
        )


@router.post("/api/spec/reorder")
async def spec_reorder(
    payload: dict = Body(...),
    user=Depends(auth.get_current_user),
):
    section_id = payload.get("sectionId")
    item_id = payload.get("itemId")
    target_parent_id = payload.get("targetParentId")
    target_index = _ensure_int(payload.get("targetIndex") or 1, 1)

    if not (section_id and item_id):
        raise HTTPException(400, "sectionId and itemId are required")

    # определяем версию для которой фиксируем порядок
    sec = await db["spec_sections"].find_one({"_id": _oid(section_id), "tenantId": _oid(user["tenantId"])})
    if not sec:
        raise HTTPException(404, "Section not found")
    for_version = int(payload.get("forVersion") or sec.get("activeVersion") or 1)

    it_before = await db["spec_items"].find_one({"_id": _oid(item_id), "tenantId": _oid(user["tenantId"])})
    if not it_before:
        raise HTTPException(404, "Item not found")
    old_parent_id = it_before.get("parentId")

    await _recalc_subtree(
        db["spec_items"],
        user["tenantId"],
        section_id,
        item_id,
        target_parent_id,
        target_index,
    )

    await _sync_version_positions_for_parent(user["tenantId"], section_id, old_parent_id, for_version, str(user.get("_id")))
    await _sync_version_positions_for_parent(user["tenantId"], section_id, target_parent_id, for_version, str(user.get("_id")))
    try:
        await log_project_action(
            db,
            user,
            project_id=sec.get("projectId"),
            action="spec.item.reorder",
            entity="spec.item",
            entity_id=str(item_id),
            message="Изменён порядок позиций спецификации",
            meta={
                **_section_meta(sec),
                "itemId": item_id,
                "targetParentId": target_parent_id,
                "targetIndex": target_index,
                "forVersion": for_version,
                "oldParentId": old_parent_id,
            },
        )
    except Exception:
        pass

    return {"ok": True}

@router.get("/api/spec/sections/{section_id}/attachments")
async def spec_section_attachments_list(
    section_id: str,
    user=Depends(auth.get_current_user),
):
    sec = await db["spec_sections"].find_one(
        {"_id": _oid(section_id), "tenantId": _oid(user["tenantId"])}
    )
    if not sec:
        raise HTTPException(404, "Section not found")

    t = _oid(user["tenantId"])
    s = sec["_id"]

    cursor = fs_spec_section.find(
        {"metadata.tenantId": t, "metadata.sectionId": s},
        sort=[("uploadDate", 1)],
    )
    files = await cursor.to_list(length=1000)

    out = []
    for f in files:
        meta = (f.get("metadata") or {}) if isinstance(f, dict) else (getattr(f, "metadata", None) or {})

        upload_dt = None
        if isinstance(f, dict):
            upload_dt = f.get("uploadDate") or f.get("upload_date")
            length = f.get("length") or 0
            fname = f.get("filename")
            fid = f.get("_id")
        else:
            upload_dt = getattr(f, "upload_date", None) or getattr(f, "uploadDate", None)
            length = getattr(f, "length", 0) or 0
            fname = getattr(f, "filename", None)
            fid = getattr(f, "_id", None)

        out.append(
            {
                "_id": str(fid),
                "filename": meta.get("originalFilename") or fname or "file",
                "size": int(length or 0),
                "contentType": meta.get("contentType") or "application/octet-stream",
                "uploadedAt": upload_dt.isoformat() if upload_dt else None,
            }
        )

    return {"items": out}

@router.post("/api/spec/sections/{section_id}/attachments")
async def spec_section_attachments_upload(
    section_id: str,
    file: UploadFile = File(...),
    user=Depends(auth.get_current_user),
):
    sec = await db["spec_sections"].find_one(
        {"_id": _oid(section_id), "tenantId": _oid(user["tenantId"])}
    )
    if not sec:
        raise HTTPException(404, "Section not found")

    now = datetime.utcnow()
    filename = (file.filename or "file").strip() or "file"

    MAX_BYTES = 25 * 1024 * 1024

    meta = {
        "tenantId": _oid(user["tenantId"]),
        "projectId": sec.get("projectId"),
        "sectionId": sec["_id"],
        "uploadedBy": str(user.get("_id")),
        "uploadedAt": now,
        "contentType": file.content_type or "application/octet-stream",
        "originalFilename": filename,
    }

    grid_in = fs_spec_section.open_upload_stream(filename, metadata=meta)

    size = 0
    try:
        while True:
            chunk = await file.read(1024 * 1024)
            if not chunk:
                break
            size += len(chunk)
            if size > MAX_BYTES:
                await grid_in.close()
                await fs_spec_section.delete(grid_in._id)
                raise HTTPException(413, "Файл слишком большой (лимит 25MB)")
            await grid_in.write(chunk)
        await grid_in.close()
    except HTTPException:
        raise
    except Exception as e:
        try:
            await fs_spec_section.delete(grid_in._id)
        except Exception:
            pass
        raise HTTPException(500, f"Upload failed: {e}")

    return {"_id": str(grid_in._id), "filename": filename, "size": size}


@router.get("/api/spec/sections/{section_id}/attachments/{file_id}/download")
async def spec_section_attachments_download(
    section_id: str,
    file_id: str,
    user=Depends(auth.get_current_user),
):
    sec = await db["spec_sections"].find_one(
        {"_id": _oid(section_id), "tenantId": _oid(user["tenantId"])}
    )
    if not sec:
        raise HTTPException(404, "Section not found")

    fid = _oid(file_id)

    try:
        grid_out = await fs_spec_section.open_download_stream(fid)
    except NoFile:
        raise HTTPException(404, "File not found")

    meta = grid_out.metadata or {}
    if meta.get("tenantId") != _oid(user["tenantId"]) or meta.get("sectionId") != sec["_id"]:
        raise HTTPException(404, "File not found")

    orig_name = meta.get("originalFilename") or grid_out.filename or "file"
    content_type = meta.get("contentType") or "application/octet-stream"

    async def iterator():
        while True:
            chunk = await grid_out.readchunk()
            if not chunk:
                break
            yield chunk

    ascii_name = _ascii_filename(orig_name)
    disp = f'attachment; filename="{ascii_name}"; filename*=UTF-8\'\'{quote(orig_name, safe="")}'

    headers = {"Content-Disposition": disp}

    return StreamingResponse(iterator(), media_type=content_type, headers=headers)

@router.delete("/api/spec/sections/{section_id}/attachments/{file_id}")
async def spec_section_attachments_delete(
    section_id: str,
    file_id: str,
    user=Depends(auth.get_current_user),
):
    sec = await db["spec_sections"].find_one(
        {"_id": _oid(section_id), "tenantId": _oid(user["tenantId"])}
    )
    if not sec:
        raise HTTPException(404, "Section not found")

    fid = _oid(file_id)

    try:
        grid_out = await fs_spec_section.open_download_stream(fid)
    except NoFile:
        raise HTTPException(404, "File not found")

    meta = grid_out.metadata or {}

    # файл должен принадлежать тому же tenant и section
    if meta.get("tenantId") != _oid(user["tenantId"]) or meta.get("sectionId") != sec["_id"]:
        raise HTTPException(404, "File not found")

    await fs_spec_section.delete(fid)

    return Response(status_code=204)

# РАБОТЫ ПО ПОЗИЦИЯМ

def _work_title(w: dict) -> str:
    return (w.get("name") or "").strip() or "Работа"


def _work_meta(w: dict) -> dict:
    return {
        "workId": str(w.get("_id")) if w.get("_id") else None,
        "projectId": str(w.get("projectId")) if w.get("projectId") else None,
        "sectionId": str(w.get("sectionId")) if w.get("sectionId") else None,
        "itemId": str(w.get("itemId")) if w.get("itemId") else None,
        "deleted": bool(w.get("deleted", False)),
    }

@router.get("/api/projects/{project_id}/spec/works")
async def spec_works_list(
    project_id: str,
    sectionId: Optional[str] = Query(None),
    itemId: Optional[str] = Query(None),
    deleted: Optional[int] = Query(None),
    user=Depends(auth.get_current_user),
):
    q: Dict[str, Any] = {
        "tenantId": _oid(user["tenantId"]),
        "projectId": _oid(project_id),
    }
    if sectionId:
        q["sectionId"] = _oid(sectionId)
    if itemId:
        q["itemId"] = _oid(itemId)
    if deleted is not None:
        q["deleted"] = bool(int(deleted))

    items = (
        await db["spec_item_works"]
        .find(q)
        .sort([("order", 1), ("createdAt", 1)])
        .to_list(100000)
    )
    return {"items": [_norm(w) for w in items]}

@router.post("/api/projects/{project_id}/spec/works")
async def spec_works_create(
    project_id: str,
    payload: dict = Body(...),
    user=Depends(auth.get_current_user),
):
    """ Создать работу по позиции спецификации """
    now = datetime.utcnow()
    item_id = payload.get("itemId")
    if not item_id:
        raise HTTPException(400, "itemId is required")

    it = await db["spec_items"].find_one(
        {"_id": _oid(item_id), "tenantId": _oid(user["tenantId"])}
    )
    if not it:
        raise HTTPException(404, "Spec item not found")

    if str(it.get("projectId")) != str(_oid(project_id)):
        raise HTTPException(400, "Item belongs to another project")

    section_id = payload.get("sectionId") or it.get("sectionId")
    if not section_id:
        raise HTTPException(400, "sectionId is required")

    name = (payload.get("name") or "").strip() or "Работа"

    last = (
        await db["spec_item_works"]
        .find(
            {
                "tenantId": _oid(user["tenantId"]),
                "projectId": _oid(project_id),
                "itemId": _oid(item_id),
            }
        )
        .sort([("order", -1)])
        .limit(1)
        .to_list(1)
    )
    next_order = int((last[0]["order"] if last else 0) or 0) + 1

    doc = {
        "tenantId": _oid(user["tenantId"]),
        "projectId": _oid(project_id),
        "sectionId": _oid(section_id),
        "itemId": _oid(item_id),
        "name": name,
        "qty_plan": _num(payload.get("qty_plan"), 0),
        "qty_fact": _num(payload.get("qty_fact"), 0),
        "order": next_order,
        "deleted": False,
        "createdAt": now,
        "updatedAt": now,
    }
    res = await db["spec_item_works"].insert_one(doc)
    created = await db["spec_item_works"].find_one({"_id": res.inserted_id})

    try:
        await log_project_action(
            db,
            user,
            project_id=project_id,
            action="spec.work.create",
            entity="spec.work",
            entity_id=str(res.inserted_id),
            message=f'Создана работа «{_work_title(created)}»',
            meta=_work_meta(created),
        )
    except Exception:
        pass

    return _norm(created)

@router.patch("/api/spec/works/{work_id}")
async def spec_works_update(
    work_id: str,
    payload: dict = Body(...),
    user=Depends(auth.get_current_user),
):
    now = datetime.utcnow()
    work = await db["spec_item_works"].find_one(
        {"_id": _oid(work_id), "tenantId": _oid(user["tenantId"])}
    )
    if not work:
        raise HTTPException(404, "Work not found")
    before = dict(work)

    updates: Dict[str, Any] = {}

    if "name" in payload:
        updates["name"] = (payload.get("name") or "").strip()
    if "qty_plan" in payload:
        updates["qty_plan"] = _num(payload.get("qty_plan"), 0)
    if "qty_fact" in payload:
        updates["qty_fact"] = _num(payload.get("qty_fact"), 0)
    if "order" in payload:
        try:
            updates["order"] = int(payload.get("order") or 0)
        except Exception:
            pass
    if "deleted" in payload:
        updates["deleted"] = bool(payload.get("deleted"))

    if "sectionId" in payload and payload.get("sectionId"):
        updates["sectionId"] = _oid(payload.get("sectionId"))
    if "itemId" in payload and payload.get("itemId"):
        updates["itemId"] = _oid(payload.get("itemId"))

    if not updates:
        return _norm(work)

    updates["updatedAt"] = now
    await db["spec_item_works"].update_one(
        {"_id": work["_id"]}, {"$set": updates}
    )

    after = await db["spec_item_works"].find_one({"_id": work["_id"]})

    try:
        diff = make_diff(before, after)
        if diff:
            await log_project_action(
                db,
                user,
                project_id=after.get("projectId"),
                action="spec.work.update",
                entity="spec.work",
                entity_id=str(after["_id"]),
                message=f'Обновлена работа «{_work_title(after)}»',
                diff=diff,
                meta=_work_meta(after),
            )
    except Exception:
        pass

    return _norm(after)

@router.delete("/api/spec/works/{work_id}")
async def spec_works_delete_forever(
    work_id: str,
    user=Depends(auth.get_current_user),
):
    work = await db["spec_item_works"].find_one(
        {"_id": _oid(work_id), "tenantId": _oid(user["tenantId"])}
    )
    if not work:
        raise HTTPException(404, "Work not found")

    await db["spec_item_works"].delete_one({"_id": work["_id"]})

    try:
        await log_project_action(
            db,
            user,
            project_id=work.get("projectId"),
            action="spec.work.delete",
            entity="spec.work",
            entity_id=str(work_id),
            message=f'Полностью удалена работа «{_work_title(work)}»',
            meta=_work_meta(work),
        )
    except Exception:
        pass

    return {"ok": True}

# ОТГРУЗКА: РАЗДЕЛЫ И ПОЗИЦИИ

def _ship_section_title(sec: dict) -> str:
    return (sec.get("title") or "").strip() or "Раздел отгрузки"


def _ship_item_title(it: dict) -> str:
    return (it.get("name") or "").strip() or "Позиция отгрузки"

@router.get("/api/projects/{project_id}/ship/sections")
async def ship_sections_list(
    project_id: str,
    deleted: Optional[int] = Query(None),
    specSectionId: Optional[str] = Query(None),
    user=Depends(auth.get_current_user),
):
    q: Dict[str, Any] = {
        "tenantId": _oid(user["tenantId"]),
        "projectId": _oid(project_id),
    }
    if deleted is not None:
        q["deleted"] = bool(int(deleted))
    if specSectionId:
        q["specSectionId"] = _oid(specSectionId)

    items = (
        await db["spec_ship_sections"]
        .find(q)
        .sort([("order", 1), ("createdAt", 1)])
        .to_list(100000)
    )
    return {"items": [_norm(x) for x in items]}

@router.post("/api/projects/{project_id}/ship/sections")
async def ship_sections_create(
    project_id: str,
    payload: dict = Body(...),
    user=Depends(auth.get_current_user),
):
    now = datetime.utcnow()
    title = (payload.get("title") or "").strip() or "Раздел отгрузки"
    spec_section_id = payload.get("specSectionId")

    spec_sec = None
    if spec_section_id:
        # проверим, что раздел спецификации существует и принадлежит проекту
        spec_sec = await db["spec_sections"].find_one(
            {"_id": _oid(spec_section_id), "tenantId": _oid(user["tenantId"])}
        )
        if not spec_sec:
            raise HTTPException(404, "Spec section not found")
        if str(spec_sec.get("projectId")) != str(_oid(project_id)):
            raise HTTPException(400, "Spec section belongs to another project")

    last = (
        await db["spec_ship_sections"]
        .find({"tenantId": _oid(user["tenantId"]), "projectId": _oid(project_id)})
        .sort([("order", -1)])
        .limit(1)
        .to_list(1)
    )
    next_order = int((last[0]["order"] if last else 0) or 0) + 1

    doc = {
        "tenantId": _oid(user["tenantId"]),
        "projectId": _oid(project_id),
        "title": title,
        "order": next_order,
        "specSectionId": _oid(spec_section_id) if spec_section_id else None,
        "comment": "",
        "deleted": False,
        "createdAt": now,
        "updatedAt": now,
    }
    res = await db["spec_ship_sections"].insert_one(doc)
    created = await db["spec_ship_sections"].find_one({"_id": res.inserted_id})

    try:
        await log_project_action(
            db,
            user,
            project_id=project_id,
            action="spec.ship.section.create",
            entity="spec.ship.section",
            entity_id=str(res.inserted_id),
            message=f'Создан раздел отгрузки «{_ship_section_title(created)}»',
            meta={
                "shipSectionId": str(res.inserted_id),
                "specSectionId": str(spec_section_id) if spec_section_id else None,
            },
        )
    except Exception:
        pass

    return _norm(created)

@router.patch("/api/ship/sections/{section_id}")
async def ship_sections_update(
    section_id: str,
    payload: dict = Body(...),
    user=Depends(auth.get_current_user),
):
    now = datetime.utcnow()
    sec = await db["spec_ship_sections"].find_one(
        {"_id": _oid(section_id), "tenantId": _oid(user["tenantId"])}
    )
    if not sec:
        raise HTTPException(404, "Shipment section not found")

    before = dict(sec)
    updates: Dict[str, Any] = {}

    if "title" in payload:
        updates["title"] = (payload.get("title") or "").strip()
    if "comment" in payload:
        updates["comment"] = payload.get("comment") or ""
    if "order" in payload:
        try:
            updates["order"] = int(payload.get("order") or 0)
        except Exception:
            pass
    if "specSectionId" in payload:
        val = payload.get("specSectionId")
        updates["specSectionId"] = _oid(val) if val else None

    cascade_deleted: Optional[bool] = None
    if "deleted" in payload:
        cascade_deleted = bool(payload.get("deleted"))
        updates["deleted"] = cascade_deleted

    if not updates:
        return _norm(sec)

    updates["updatedAt"] = now
    await db["spec_ship_sections"].update_one(
        {"_id": sec["_id"]}, {"$set": updates}
    )

    if cascade_deleted is not None:
        now2 = datetime.utcnow()
        await db["spec_ship_items"].update_many(
            {
                "tenantId": _oid(user["tenantId"]),
                "shipmentSectionId": sec["_id"],
            },
            {"$set": {"deleted": cascade_deleted, "updatedAt": now2}},
        )

    after = await db["spec_ship_sections"].find_one({"_id": sec["_id"]})

    try:
        diff = make_diff(before, after)
        if diff:
            await log_project_action(
                db,
                user,
                project_id=after.get("projectId"),
                action="spec.ship.section.update",
                entity="spec.ship.section",
                entity_id=str(after["_id"]),
                message=f'Обновлён раздел отгрузки «{_ship_section_title(after)}»',
                diff=diff,
                meta={
                    "shipSectionId": str(after["_id"]),
                    "specSectionId": str(after.get("specSectionId")),
                },
            )
    except Exception:
        pass

    return _norm(after)

@router.delete("/api/ship/sections/{section_id}")
async def ship_sections_delete_forever(
    section_id: str,
    user=Depends(auth.get_current_user),
):
    sec = await db["spec_ship_sections"].find_one(
        {"_id": _oid(section_id), "tenantId": _oid(user["tenantId"])}
    )
    if not sec:
        raise HTTPException(404, "Shipment section not found")

    res_items = await db["spec_ship_items"].delete_many(
        {
            "tenantId": _oid(user["tenantId"]),
            "shipmentSectionId": sec["_id"],
        }
    )
    await db["spec_ship_sections"].delete_one({"_id": sec["_id"]})

    try:
        await log_project_action(
            db,
            user,
            project_id=sec.get("projectId"),
            action="spec.ship.section.delete",
            entity="spec.ship.section",
            entity_id=str(section_id),
            message=f'Полностью удалён раздел отгрузки «{_ship_section_title(sec)}»',
            meta={
                "shipSectionId": str(section_id),
                "specSectionId": str(sec.get("specSectionId")),
                "itemsDeleted": res_items.deleted_count,
            },
        )
    except Exception:
        pass

    return {"ok": True}

@router.get("/api/projects/{project_id}/ship/items")
async def ship_items_list(
    project_id: str,
    sectionId: Optional[str] = Query(None),
    shipmentSectionId: Optional[str] = Query(None),
    specItemId: Optional[str] = Query(None),
    specSectionId: Optional[str] = Query(None),
    deleted: Optional[int] = Query(None),
    user=Depends(auth.get_current_user),
):
    q: Dict[str, Any] = {
        "tenantId": _oid(user["tenantId"]),
        "projectId": _oid(project_id),
    }

    sid = sectionId or shipmentSectionId
    if sid:
        q["shipmentSectionId"] = _oid(sid)
    if specItemId:
        q["specItemId"] = _oid(specItemId)
    if specSectionId:
        q["specSectionId"] = _oid(specSectionId)
    if deleted is not None:
        q["deleted"] = bool(int(deleted))

    items = (
        await db["spec_ship_items"]
        .find(q)
        .sort([("pos", 1), ("createdAt", 1)])
        .to_list(100000)
    )
    return {"items": [_norm(x) for x in items]}

@router.post("/api/projects/{project_id}/ship/items")
async def ship_items_create(
    project_id: str,
    payload: dict = Body(...),
    user=Depends(auth.get_current_user),
):
    now = datetime.utcnow()
    section_id = payload.get("shipmentSectionId") or payload.get("sectionId")
    if not section_id:
        raise HTTPException(400, "shipmentSectionId is required")

    sec = await db["spec_ship_sections"].find_one(
        {"_id": _oid(section_id), "tenantId": _oid(user["tenantId"])}
    )
    if not sec:
        raise HTTPException(404, "Shipment section not found")
    if str(sec.get("projectId")) != str(_oid(project_id)):
        raise HTTPException(400, "Shipment section belongs to another project")

    name = (payload.get("name") or "").strip() or "Позиция отгрузки"

    last = (
        await db["spec_ship_items"]
        .find(
            {
                "tenantId": _oid(user["tenantId"]),
                "projectId": _oid(project_id),
                "shipmentSectionId": sec["_id"],
            }
        )
        .sort([("pos", -1)])
        .limit(1)
        .to_list(1)
    )
    next_pos = int((last[0]["pos"] if last else 0) or 0) + 1

    spec_item_id = payload.get("specItemId")
    spec_section_id = payload.get("specSectionId")

    doc = {
        "tenantId": _oid(user["tenantId"]),
        "projectId": _oid(project_id),
        "shipmentSectionId": sec["_id"],
        "specItemId": _oid(spec_item_id) if spec_item_id else None,
        "specSectionId": _oid(spec_section_id) if spec_section_id else None,
        "pos": next_pos,
        "name": name,
        "unit": (payload.get("unit") or "").strip() or None,
        "qty": _num(payload.get("qty"), 0),
        "price": _num(payload.get("price"), 0),
        "deleted": False,
        "createdAt": now,
        "updatedAt": now,
    }
    res = await db["spec_ship_items"].insert_one(doc)
    created = await db["spec_ship_items"].find_one({"_id": res.inserted_id})

    try:
        await log_project_action(
            db,
            user,
            project_id=project_id,
            action="spec.ship.item.create",
            entity="spec.ship.item",
            entity_id=str(res.inserted_id),
            message=f'Создана позиция отгрузки «{_ship_item_title(created)}»',
            meta={
                "shipItemId": str(res.inserted_id),
                "shipSectionId": str(sec["_id"]),
                "specItemId": str(spec_item_id) if spec_item_id else None,
                "specSectionId": str(spec_section_id) if spec_section_id else None,
            },
        )
    except Exception:
        pass

    return _norm(created)

@router.patch("/api/ship/items/{item_id}")
async def ship_items_update(
    item_id: str,
    payload: dict = Body(...),
    user=Depends(auth.get_current_user),
):
    now = datetime.utcnow()
    it = await db["spec_ship_items"].find_one(
        {"_id": _oid(item_id), "tenantId": _oid(user["tenantId"])}
    )
    if not it:
        raise HTTPException(404, "Shipment item not found")

    before = dict(it)
    updates: Dict[str, Any] = {}

    if "name" in payload:
        updates["name"] = (payload.get("name") or "").strip()
    if "unit" in payload:
        updates["unit"] = (payload.get("unit") or "").strip() or None
    if "qty" in payload:
        updates["qty"] = _num(payload.get("qty"), 0)
    if "price" in payload:
        updates["price"] = _num(payload.get("price"), 0)
    if "pos" in payload:
        try:
            updates["pos"] = int(payload.get("pos") or 0)
        except Exception:
            pass
    if "specItemId" in payload:
        val = payload.get("specItemId")
        updates["specItemId"] = _oid(val) if val else None
    if "specSectionId" in payload:
        val = payload.get("specSectionId")
        updates["specSectionId"] = _oid(val) if val else None
    if "shipmentSectionId" in payload or "sectionId" in payload:
        sid = payload.get("shipmentSectionId") or payload.get("sectionId")
        if sid:
            updates["shipmentSectionId"] = _oid(sid)
    if "deleted" in payload:
        updates["deleted"] = bool(payload.get("deleted"))

    if not updates:
        return _norm(it)

    updates["updatedAt"] = now
    await db["spec_ship_items"].update_one(
        {"_id": it["_id"]}, {"$set": updates}
    )

    after = await db["spec_ship_items"].find_one({"_id": it["_id"]})

    try:
        diff = make_diff(before, after)
        if diff:
            await log_project_action(
                db,
                user,
                project_id=after.get("projectId"),
                action="spec.ship.item.update",
                entity="spec.ship.item",
                entity_id=str(after["_id"]),
                message=f'Обновлена позиция отгрузки «{_ship_item_title(after)}»',
                diff=diff,
                meta={
                    "shipItemId": str(after["_id"]),

                    "shipSectionId": str(after.get("shipmentSectionId")),
                    "specItemId": str(after.get("specItemId")),
                    "specSectionId": str(after.get("specSectionId")),
                },
            )
    except Exception:
        pass

    return _norm(after)

@router.delete("/api/ship/items/{item_id}")
async def ship_items_delete_forever(
    item_id: str,
    user=Depends(auth.get_current_user),
):
    it = await db["spec_ship_items"].find_one(
        {"_id": _oid(item_id), "tenantId": _oid(user["tenantId"])}
    )
    if not it:
        raise HTTPException(404, "Shipment item not found")

    await db["spec_ship_items"].delete_one({"_id": it["_id"]})

    try:
        await log_project_action(
            db,
            user,
            project_id=it.get("projectId"),
            action="spec.ship.item.delete",
            entity="spec.ship.item",
            entity_id=str(item_id),
            message=f'Полностью удалена позиция отгрузки «{_ship_item_title(it)}»',
            meta={
                "shipItemId": str(item_id),
                "shipSectionId": str(it.get("shipmentSectionId")),
                "specItemId": str(it.get("specItemId")),
                "specSectionId": str(it.get("specSectionId")),
            },
        )
    except Exception:
        pass

    return {"ok": True}

@router.get("/api/ship/sections/{section_id}/export")
async def ship_section_export_excel(
    section_id: str,
    user=Depends(auth.get_current_user),
):
    """ экспорт отгрузки в excel """

    sec = await db["spec_ship_sections"].find_one(
        {"_id": _oid(section_id), "tenantId": _oid(user["tenantId"])}
    )
    if not sec:
        raise HTTPException(404, "Shipment section not found")

    tenant_id = _oid(user["tenantId"])
    section_oid = sec["_id"]
    project_id = sec.get("projectId")

    items = (
        await db["spec_ship_items"]
        .find(
            {
                "tenantId": tenant_id,
                "projectId": project_id,
                "shipmentSectionId": section_oid,
                "deleted": {"$ne": True},
            }
        )
        .sort([("pos", 1), ("createdAt", 1)])
        .to_list(100000)
    )

    # Подтягиваем названия разделов и позиций спецификации, если есть связи

    spec_section_titles: dict = {}
    spec_section_ids = {
        it.get("specSectionId")
        for it in items
        if it.get("specSectionId")
    }
    if spec_section_ids:
        cur = db["spec_sections"].find(
            {
                "_id": {"$in": list(spec_section_ids)},
                "tenantId": tenant_id,
            }
        )
        secs = await cur.to_list(100000)
        for s in secs:
            title = s.get("title") or s.get("name") or ""
            ver = s.get("versionLabel") or s.get("version")
            if ver:
                spec_section_titles[s["_id"]] = f"{title} · {ver}"
            else:
                spec_section_titles[s["_id"]] = title

    spec_item_docs: dict = {}
    spec_item_ids = {
        it.get("specItemId")
        for it in items
        if it.get("specItemId")
    }
    if spec_item_ids:
        cur = db["spec_items"].find(
            {
                "_id": {"$in": list(spec_item_ids)},
                "tenantId": tenant_id,
            }
        )
        spec_items = await cur.to_list(100000)
        for si in spec_items:
            spec_item_docs[si["_id"]] = si

    def _spec_item_label_for_ship(it: dict) -> str:

        for key in ("specItemLabel", "specItemTitle", "specItemName", "specName"):
            if it.get(key):
                return str(it[key])

        # Подтягиваем исходный документ spec_items
        sid = it.get("specItemId")
        if sid and sid in spec_item_docs:
            si = spec_item_docs[sid]
            num = si.get("number") or si.get("posHuman") or si.get("pos")
            name = si.get("name") or si.get("title") or ""
            if num not in (None, "", 0):
                return f"{num}. {name}"
            return name

        # Если ничего нет - пробуем хотя бы раздел
        sec_id = it.get("specSectionId")
        if sec_id and sec_id in spec_section_titles:
            return spec_section_titles[sec_id]

        return ""

    # Готовим excel

    from openpyxl import Workbook
    from openpyxl.styles import Border, Side, Alignment, Font
    from openpyxl.utils import get_column_letter
    from io import BytesIO

    wb = Workbook()
    ws = wb.active
    ws.title = "Отгрузка"

    headers = ["№", "Наименование", "Раздел", "Количество", "Цена", "Сумма"]
    ws.append(headers)

    thin = Side(style="thin", color="000000")
    border_all = Border(left=thin, right=thin, top=thin, bottom=thin)

    for idx, it in enumerate(items, start=1):
        qty = it.get("qty") or 0
        price = it.get("price") or 0
        total = float(qty) * float(price)

        # имя из отгрузки
        base_name = (it.get("name") or "").strip()

        spec_label = _spec_item_label_for_ship(it)
        if spec_label:
            name_cell = f"{base_name} ({spec_label})"
        else:
            name_cell = base_name

        # название раздела для отдельной колонки
        section_title = ""
        sec_id = it.get("specSectionId")
        if sec_id and sec_id in spec_section_titles:
            section_title = spec_section_titles[sec_id]

        ws.append(
            [
                idx,
                name_cell,
                section_title,
                qty,
                price,
                total,
            ]
        )

    # Ширина колонок
    widths = {
        1: 5,   # №
        2: 60,  # Наименование
        3: 30,  # Раздел
        4: 15,  # Количество
        5: 15,  # Цена
        6: 18,  # Сумма
    }
    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    max_row = ws.max_row
    max_col = ws.max_column

    for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
        for cell in row:
            cell.border = border_all
            if cell.row == 1:
                continue
            if cell.column == 1:
                horiz = "center"
            elif cell.column in (4, 5, 6):
                horiz = "right"
            else:
                horiz = "left"
            cell.alignment = Alignment(
                wrap_text=True,
                vertical="top",
                horizontal=horiz,
            )

    # шапка
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )

    # файл в поток
    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)

    safe_title = _ship_section_title(sec).replace('"', "'").replace("\n", " ")
    ascii_title = re.sub(r"[^A-Za-z0-9_.-]+", "_", safe_title).strip("_") or "shipment"
    filename = f"{ascii_title}.xlsx"
    headers_resp = {"Content-Disposition": f'attachment; filename="{filename}"'}

    try:
        await log_project_action(
            db,
            user,
            project_id=project_id,
            action="spec.ship.section.export_excel",
            entity="spec.ship.section",
            entity_id=str(sec["_id"]),
            message=f'Экспортирована отгрузка «{_ship_section_title(sec)}» в Excel',
            meta={"shipSectionId": str(sec["_id"])},
        )
    except Exception:
        pass

    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers_resp,
    )

@router.post("/api/ship/sections/{section_id}/import")
async def ship_section_import_excel(
    section_id: str,
    file: UploadFile = File(...),
    user=Depends(auth.get_current_user),
):
    """ Импорт позиций отгрузки из excel-накладной  """
    sec = await db["spec_ship_sections"].find_one(
        {"_id": _oid(section_id), "tenantId": _oid(user["tenantId"])}
    )
    if not sec:
        raise HTTPException(404, "Shipment section not found")

    filename = file.filename or ""
    if not filename.lower().endswith(".xlsx"):
        raise HTTPException(400, "Ожидается .xlsx файл")

    try:
        content = await file.read()
        if not content:
            raise HTTPException(400, "Пустой файл")

        stream = BytesIO(content)
        wb = load_workbook(stream, data_only=True)
        ws = wb.active
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(400, f"Не удалось прочитать Excel: {e}")

    if ws.max_row < 2:
        raise HTTPException(400, "В файле нет данных")

    # Наименование
    name_col, name_hdr_row = _find_excel_col(
        ws,
        ["наимен", "материальн", "материальные ценности"],
    )
    # Кол-во "отпущено"
    qty_col, qty_hdr_row = _find_excel_col(
        ws,
        ["отпущ", "отпущено"],
    )
    # Сумма с НДС
    sum_vat_col, sum_vat_hdr_row = _find_excel_col(
        ws,
        ["с учетом ндс", "с учётом ндс", "сумма с учетом ндс", "сумма с учётом ндс"],
    )
    # Цена за единицу
    price_col, price_hdr_row = _find_excel_col(
        ws,
        ["цена", "цена, руб", "цена руб"],
    )

    if not name_col:
        raise HTTPException(400, "Не найдена колонка с наименованием")
    if not qty_col:
        raise HTTPException(400, "Не найдена колонка количества «отпущено»")
    if not sum_vat_col:
        raise HTTPException(400, "Не найдена колонка «Сумма с учетом НДС»")

    header_rows = [
        r
        for r in [name_hdr_row, qty_hdr_row, sum_vat_hdr_row, price_hdr_row]
        if r
    ]
    start_row = max(header_rows) + 1 if header_rows else 2

    def is_row_empty(row_idx: int) -> bool:
        for col in range(1, ws.max_column + 1):
            if ws.cell(row=row_idx, column=col).value not in (None, "", " "):
                return False
        return True

    tenant_id = _oid(user["tenantId"])
    project_id = sec.get("projectId")
    section_oid = sec["_id"]

    # стартовый pos
    last = (
        await db["spec_ship_items"]
        .find(
            {
                "tenantId": tenant_id,
                "projectId": project_id,
                "shipmentSectionId": section_oid,
            }
        )
        .sort([("pos", -1)])
        .limit(1)
        .to_list(1)
    )
    pos_counter = int((last[0]["pos"] if last else 0) or 0)

    now = datetime.utcnow()
    docs_to_insert: List[dict] = []

    for row in range(start_row, ws.max_row + 1):
        if is_row_empty(row):
            continue

        name_val = ws.cell(row=row, column=name_col).value
        qty_val = ws.cell(row=row, column=qty_col).value
        price_unit_val = ws.cell(row=row, column=price_col).value if price_col else None
        sum_vat_val = ws.cell(row=row, column=sum_vat_col).value if sum_vat_col else None

        name = (str(name_val).strip() if name_val is not None else "")
        if not name:
            continue

        # пропускаем строки "Итого", "Всего" и т.п
        name_lower = name.lower().replace("ё", "е")
        if name_lower.startswith("итого") or name_lower.startswith("всего"):
            continue

        qty = _num_ru(qty_val, 0.0)
        if qty <= 0:
            continue

        # ЦЕНА В СИСТЕМЕ = СУММА С УЧЁТОМ НДС

        price = 0.0
        if sum_vat_val is not None and str(sum_vat_val).strip() != "":
            total_vat = _num_ru(sum_vat_val, 0.0)
            if total_vat > 0:
                price = total_vat

        if price <= 0 and price_unit_val is not None:
            unit_price = _num_ru(price_unit_val, 0.0)
            if unit_price > 0:
                price = unit_price * qty if qty > 0 else unit_price

        pos_counter += 1
        doc = {
            "tenantId": tenant_id,
            "projectId": project_id,
            "shipmentSectionId": section_oid,
            "specItemId": None,
            "specSectionId": None,
            "pos": pos_counter,
            "name": name,
            "unit": None,
            "qty": qty,
            "price": price,
            "deleted": False,
            "createdAt": now,
            "updatedAt": now,
        }
        docs_to_insert.append(doc)

    if not docs_to_insert:
        raise HTTPException(400, "Не удалось найти ни одной позиции для импорта")

    await db["spec_ship_items"].insert_many(docs_to_insert)

    try:
        await log_project_action(
            db,
            user,
            project_id=project_id,
            action="spec.ship.section.import_excel",
            entity="spec.ship.section",
            entity_id=str(sec["_id"]),
            message=f'Импортирована отгрузка «{_ship_section_title(sec)}» из Excel',
            meta={
                "shipSectionId": str(sec["_id"]),
                "itemsImported": len(docs_to_insert),
            },
        )
    except Exception:
        pass

    after = await db["spec_ship_sections"].find_one({"_id": sec["_id"]})

    return {
        "section": _norm(after),
        "itemsImported": len(docs_to_insert),
    }

# Вложения для отгрузок (накладных)
@router.get("/api/ship/sections/{section_id}/attachments")
async def ship_section_attachments_list(
    section_id: str,
    user=Depends(auth.get_current_user),
):
    sec = await db["spec_ship_sections"].find_one(
        {"_id": _oid(section_id), "tenantId": _oid(user["tenantId"])}
    )
    if not sec:
        raise HTTPException(404, "Shipment section not found")

    t = _oid(user["tenantId"])
    s = sec["_id"]

    cursor = fs_ship_section.find(
        {"metadata.tenantId": t, "metadata.sectionId": s},
        sort=[("uploadDate", 1)],
    )
    files = await cursor.to_list(length=1000)

    out = []
    for f in files:
        meta = (f.get("metadata") or {}) if isinstance(f, dict) else (getattr(f, "metadata", None) or {})

        upload_dt = None
        if isinstance(f, dict):
            upload_dt = f.get("uploadDate") or f.get("upload_date")
            length = f.get("length") or 0
            fname = f.get("filename")
            fid = f.get("_id")
        else:
            upload_dt = getattr(f, "upload_date", None) or getattr(f, "uploadDate", None)
            length = getattr(f, "length", 0) or 0
            fname = getattr(f, "filename", None)
            fid = getattr(f, "_id", None)

        out.append(
            {
                "_id": str(fid),
                "filename": meta.get("originalFilename") or fname or "file",
                "size": int(length or 0),
                "contentType": meta.get("contentType") or "application/octet-stream",
                "uploadedAt": upload_dt.isoformat() if upload_dt else None,
            }
        )

    return {"items": out}

@router.post("/api/ship/sections/{section_id}/attachments")
async def ship_section_attachments_upload(
    section_id: str,
    file: UploadFile = File(...),
    user=Depends(auth.get_current_user),
):
    sec = await db["spec_ship_sections"].find_one(
        {"_id": _oid(section_id), "tenantId": _oid(user["tenantId"])}
    )
    if not sec:
        raise HTTPException(404, "Shipment section not found")

    now = datetime.utcnow()
    filename = (file.filename or "file").strip() or "file"

    MAX_BYTES = 25 * 1024 * 1024

    meta = {
        "tenantId": _oid(user["tenantId"]),
        "projectId": sec.get("projectId"),
        "sectionId": sec["_id"],
        "uploadedBy": str(user.get("_id")),
        "uploadedAt": now,
        "contentType": file.content_type or "application/octet-stream",
        "originalFilename": filename,
    }

    grid_in = fs_ship_section.open_upload_stream(filename, metadata=meta)

    size = 0
    try:
        while True:
            chunk = await file.read(1024 * 1024)
            if not chunk:
                break
            size += len(chunk)
            if size > MAX_BYTES:
                await grid_in.close()
                await fs_ship_section.delete(grid_in._id)
                raise HTTPException(413, "Файл слишком большой (лимит 25MB)")
            await grid_in.write(chunk)
        await grid_in.close()
    except HTTPException:
        raise
    except Exception as e:
        try:
            await fs_ship_section.delete(grid_in._id)
        except Exception:
            pass
        raise HTTPException(500, f"Upload failed: {e}")

    return {"_id": str(grid_in._id), "filename": filename, "size": size}


@router.get("/api/ship/sections/{section_id}/attachments/{file_id}/download")
async def ship_section_attachments_download(
    section_id: str,
    file_id: str,
    user=Depends(auth.get_current_user),
):
    sec = await db["spec_ship_sections"].find_one(
        {"_id": _oid(section_id), "tenantId": _oid(user["tenantId"])}
    )
    if not sec:
        raise HTTPException(404, "Shipment section not found")

    fid = _oid(file_id)

    try:
        grid_out = await fs_ship_section.open_download_stream(fid)
    except NoFile:
        raise HTTPException(404, "File not found")

    meta = grid_out.metadata or {}
    if meta.get("tenantId") != _oid(user["tenantId"]) or meta.get("sectionId") != sec["_id"]:
        raise HTTPException(404, "File not found")

    orig_name = meta.get("originalFilename") or grid_out.filename or "file"
    content_type = meta.get("contentType") or "application/octet-stream"

    async def iterator():
        while True:
            chunk = await grid_out.readchunk()
            if not chunk:
                break
            yield chunk

    ascii_name = _ascii_filename(orig_name)
    disp = f'attachment; filename="{ascii_name}"; filename*=UTF-8\'\'{quote(orig_name, safe="")}'
    headers = {"Content-Disposition": disp}

    return StreamingResponse(iterator(), media_type=content_type, headers=headers)

@router.delete("/api/ship/sections/{section_id}/attachments/{file_id}")
async def ship_section_attachments_delete(
    section_id: str,
    file_id: str,
    user=Depends(auth.get_current_user),
):
    sec = await db["spec_ship_sections"].find_one(
        {"_id": _oid(section_id), "tenantId": _oid(user["tenantId"])}
    )
    if not sec:
        raise HTTPException(404, "Shipment section not found")

    fid = _oid(file_id)

    try:
        grid_out = await fs_ship_section.open_download_stream(fid)
    except NoFile:
        raise HTTPException(404, "File not found")

    meta = grid_out.metadata or {}
    if meta.get("tenantId") != _oid(user["tenantId"]) or meta.get("sectionId") != sec["_id"]:
        raise HTTPException(404, "File not found")

    await fs_ship_section.delete(fid)
    return Response(status_code=204)

# ВЫПОЛНЕНИЕ: РАЗДЕЛЫ И ПОЗИЦИИ

def _exec_section_title(sec: dict) -> str:
    return (sec.get("title") or "").strip() or "Раздел выполнения"


def _exec_item_title(it: dict) -> str:
    return (it.get("name") or "").strip() or "Позиция выполнения"


@router.get("/api/projects/{project_id}/exec/sections")
async def exec_sections_list(
    project_id: str,
    deleted: Optional[int] = Query(None),
    specSectionId: Optional[str] = Query(None),
    user=Depends(auth.get_current_user),
):
    q: Dict[str, Any] = {
        "tenantId": _oid(user["tenantId"]),
        "projectId": _oid(project_id),
    }
    if deleted is not None:
        q["deleted"] = bool(int(deleted))
    if specSectionId:
        q["specSectionId"] = _oid(specSectionId)

    items = (
        await db["spec_exec_sections"]
        .find(q)
        .sort([("order", 1), ("createdAt", 1)])
        .to_list(100000)
    )
    return {"items": [_norm(x) for x in items]}


@router.post("/api/projects/{project_id}/exec/sections")
async def exec_sections_create(
    project_id: str,
    payload: dict = Body(...),
    user=Depends(auth.get_current_user),
):
    now = datetime.utcnow()
    title = (payload.get("title") or "").strip() or "Раздел выполнения"
    spec_section_id = payload.get("specSectionId")

    if not spec_section_id:
        raise HTTPException(400, "specSectionId is required")

    spec_sec = await db["spec_sections"].find_one(
        {"_id": _oid(spec_section_id), "tenantId": _oid(user["tenantId"])}
    )
    if not spec_sec:
        raise HTTPException(404, "Spec section not found")
    if str(spec_sec.get("projectId")) != str(_oid(project_id)):
        raise HTTPException(400, "Spec section belongs to another project")

    last = (
        await db["spec_exec_sections"]
        .find({"tenantId": _oid(user["tenantId"]), "projectId": _oid(project_id)})
        .sort([("order", -1)])
        .limit(1)
        .to_list(1)
    )
    next_order = int((last[0]["order"] if last else 0) or 0) + 1

    doc = {
        "tenantId": _oid(user["tenantId"]),
        "projectId": _oid(project_id),
        "title": title,
        "order": next_order,
        "specSectionId": _oid(spec_section_id),
        "deleted": False,
        "createdAt": now,
        "updatedAt": now,
    }
    res = await db["spec_exec_sections"].insert_one(doc)
    created = await db["spec_exec_sections"].find_one({"_id": res.inserted_id})

    try:
        await log_project_action(
            db,
            user,
            project_id=project_id,
            action="spec.exec.section.create",
            entity="spec.exec.section",
            entity_id=str(res.inserted_id),
            message=f'Создан раздел выполнения «{_exec_section_title(created)}»',
            meta={
                "execSectionId": str(res.inserted_id),
                "specSectionId": str(spec_section_id),
            },
        )
    except Exception:
        pass

    return _norm(created)

@router.patch("/api/exec/sections/{section_id}")
async def exec_sections_update(
    section_id: str,
    payload: dict = Body(...),
    user=Depends(auth.get_current_user),
):
    now = datetime.utcnow()
    sec = await db["spec_exec_sections"].find_one(
        {"_id": _oid(section_id), "tenantId": _oid(user["tenantId"])}
    )
    if not sec:
        raise HTTPException(404, "Exec section not found")

    before = dict(sec)
    updates: Dict[str, Any] = {}

    if "title" in payload:
        updates["title"] = (payload.get("title") or "").strip()
    if "order" in payload:
        try:
            updates["order"] = int(payload.get("order") or 0)
        except Exception:
            pass
    if "specSectionId" in payload and payload.get("specSectionId"):
        updates["specSectionId"] = _oid(payload.get("specSectionId"))
    cascade_deleted: Optional[bool] = None
    if "deleted" in payload:
        cascade_deleted = bool(payload.get("deleted"))
        updates["deleted"] = cascade_deleted

    if not updates:
        return _norm(sec)

    updates["updatedAt"] = now
    await db["spec_exec_sections"].update_one(
        {"_id": sec["_id"]}, {"$set": updates}
    )

    if cascade_deleted is not None:
        now2 = datetime.utcnow()
        await db["spec_exec_items"].update_many(
            {
                "tenantId": _oid(user["tenantId"]),
                "execSectionId": sec["_id"],
            },
            {"$set": {"deleted": cascade_deleted, "updatedAt": now2}},
        )

    after = await db["spec_exec_sections"].find_one({"_id": sec["_id"]})

    try:
        diff = make_diff(before, after)
        if diff:
            await log_project_action(
                db,
                user,
                project_id=after.get("projectId"),
                action="spec.exec.section.update",
                entity="spec.exec.section",
                entity_id=str(after["_id"]),
                message=f'Обновлён раздел выполнения «{_exec_section_title(after)}»',
                diff=diff,
                meta={
                    "execSectionId": str(after["_id"]),
                    "specSectionId": str(after.get("specSectionId")),
                },
            )
    except Exception:
        pass

    return _norm(after)

@router.delete("/api/exec/sections/{section_id}")
async def exec_sections_delete_forever(
    section_id: str,
    user=Depends(auth.get_current_user),
):
    sec = await db["spec_exec_sections"].find_one(
        {"_id": _oid(section_id), "tenantId": _oid(user["tenantId"])}
    )
    if not sec:
        raise HTTPException(404, "Exec section not found")

    res_items = await db["spec_exec_items"].delete_many(
        {
            "tenantId": _oid(user["tenantId"]),
            "execSectionId": sec["_id"],
        }
    )
    await db["spec_exec_sections"].delete_one({"_id": sec["_id"]})

    try:
        await log_project_action(
            db,
            user,
            project_id=sec.get("projectId"),
            action="spec.exec.section.delete",
            entity="spec.exec.section",
            entity_id=str(section_id),
            message=f'Полностью удалён раздел выполнения «{_exec_section_title(sec)}»',
            meta={
                "execSectionId": str(section_id),
                "specSectionId": str(sec.get("specSectionId")),
                "itemsDeleted": res_items.deleted_count,
            },
        )
    except Exception:
        pass

    return {"ok": True}

@router.get("/api/projects/{project_id}/exec/items")
async def exec_items_list(
    project_id: str,
    sectionId: Optional[str] = Query(None),
    execSectionId: Optional[str] = Query(None),
    specItemId: Optional[str] = Query(None),
    deleted: Optional[int] = Query(None),
    user=Depends(auth.get_current_user),
):
    q: Dict[str, Any] = {
        "tenantId": _oid(user["tenantId"]),
        "projectId": _oid(project_id),
    }

    sid = sectionId or execSectionId
    if sid:
        q["execSectionId"] = _oid(sid)
    if specItemId:
        q["specItemId"] = _oid(specItemId)
    if deleted is not None:
        q["deleted"] = bool(int(deleted))

    items = (
        await db["spec_exec_items"]
        .find(q)
        .sort([("pos", 1), ("createdAt", 1)])
        .to_list(100000)
    )
    return {"items": [_norm(x) for x in items]}

@router.post("/api/projects/{project_id}/exec/items")
async def exec_items_create(
    project_id: str,
    payload: dict = Body(...),
    user=Depends(auth.get_current_user),
):
    now = datetime.utcnow()
    section_id = payload.get("execSectionId") or payload.get("sectionId")
    if not section_id:
        raise HTTPException(400, "execSectionId is required")

    sec = await db["spec_exec_sections"].find_one(
        {"_id": _oid(section_id), "tenantId": _oid(user["tenantId"])}
    )
    if not sec:
        raise HTTPException(404, "Exec section not found")
    if str(sec.get("projectId")) != str(_oid(project_id)):
        raise HTTPException(400, "Exec section belongs to another project")

    name = (payload.get("name") or "").strip() or "Позиция выполнения"

    last = (
        await db["spec_exec_items"]
        .find(
            {
                "tenantId": _oid(user["tenantId"]),
                "projectId": _oid(project_id),
                "execSectionId": sec["_id"],
            }
        )
        .sort([("pos", -1)])
        .limit(1)
        .to_list(1)
    )
    next_pos = int((last[0]["pos"] if last else 0) or 0) + 1

    spec_item_id = payload.get("specItemId")

    doc = {
        "tenantId": _oid(user["tenantId"]),
        "projectId": _oid(project_id),
        "execSectionId": sec["_id"],
        "specItemId": _oid(spec_item_id) if spec_item_id else None,
        "pos": next_pos,
        "name": name,
        "unit": (payload.get("unit") or "").strip() or None,
        "qty": _num(payload.get("qty"), 0),  # факт
        "deleted": False,
        "createdAt": now,
        "updatedAt": now,
    }
    res = await db["spec_exec_items"].insert_one(doc)
    created = await db["spec_exec_items"].find_one({"_id": res.inserted_id})

    try:
        await log_project_action(
            db,
            user,
            project_id=project_id,
            action="spec.exec.item.create",
            entity="spec.exec.item",
            entity_id=str(res.inserted_id),
            message=f'Создана позиция выполнения «{_exec_item_title(created)}»',
            meta={
                "execItemId": str(res.inserted_id),
                "execSectionId": str(sec["_id"]),
                "specItemId": str(spec_item_id) if spec_item_id else None,
            },
        )
    except Exception:
        pass

    return _norm(created)

@router.patch("/api/exec/items/{item_id}")
async def exec_items_update(
    item_id: str,
    payload: dict = Body(...),
    user=Depends(auth.get_current_user),
):
    now = datetime.utcnow()
    it = await db["spec_exec_items"].find_one(
        {"_id": _oid(item_id), "tenantId": _oid(user["tenantId"])}
    )
    if not it:
        raise HTTPException(404, "Exec item not found")

    before = dict(it)
    updates: Dict[str, Any] = {}

    if "name" in payload:
        updates["name"] = (payload.get("name") or "").strip()
    if "unit" in payload:
        updates["unit"] = (payload.get("unit") or "").strip() or None
    if "qty" in payload:
        updates["qty"] = _num(payload.get("qty"), 0)
    if "pos" in payload:
        try:
            updates["pos"] = int(payload.get("pos") or 0)
        except Exception:
            pass
    if "specItemId" in payload:
        val = payload.get("specItemId")
        updates["specItemId"] = _oid(val) if val else None
    if "execSectionId" in payload or "sectionId" in payload:
        sid = payload.get("execSectionId") or payload.get("sectionId")
        if sid:
            updates["execSectionId"] = _oid(sid)
    if "deleted" in payload:
        updates["deleted"] = bool(payload.get("deleted"))

    if not updates:
        return _norm(it)

    updates["updatedAt"] = now
    await db["spec_exec_items"].update_one(
        {"_id": it["_id"]}, {"$set": updates}
    )

    after = await db["spec_exec_items"].find_one({"_id": it["_id"]})

    try:
        diff = make_diff(before, after)
        if diff:
            await log_project_action(
                db,
                user,
                project_id=after.get("projectId"),
                action="spec.exec.item.update",
                entity="spec.exec.item",
                entity_id=str(after["_id"]),
                message=f'Обновлена позиция выполнения «{_exec_item_title(after)}»',
                diff=diff,
                meta={
                    "execItemId": str(after["_id"]),
                    "execSectionId": str(after.get("execSectionId")),
                    "specItemId": str(after.get("specItemId")),
                },
            )
    except Exception:
        pass

    return _norm(after)

@router.delete("/api/exec/items/{item_id}")
async def exec_items_delete_forever(
    item_id: str,
    user=Depends(auth.get_current_user),
):
    it = await db["spec_exec_items"].find_one(
        {"_id": _oid(item_id), "tenantId": _oid(user["tenantId"])}
    )
    if not it:
        raise HTTPException(404, "Exec item not found")

    await db["spec_exec_items"].delete_one({"_id": it["_id"]})

    try:
        await log_project_action(
            db,
            user,
            project_id=it.get("projectId"),
            action="spec.exec.item.delete",
            entity="spec.exec.item",
            entity_id=str(item_id),
            message=f'Полностью удалена позиция выполнения «{_exec_item_title(it)}»',
            meta={
                "execItemId": str(item_id),
                "execSectionId": str(it.get("execSectionId")),
                "specItemId": str(it.get("specItemId")),
            },
        )
    except Exception:
        pass

    return {"ok": True}

@router.get("/api/spec/summary/{section_id}/export")
async def spec_summary_export_excel(
    section_id: str,
    user=Depends(auth.get_current_user),
):
    """ Экспорт свода по разделу спецификации в Excel """
    tenant_id = _oid(user["tenantId"])

    sec = await db["spec_sections"].find_one(
        {"_id": _oid(section_id), "tenantId": tenant_id}
    )
    if not sec:
        raise HTTPException(404, "Spec section not found")

    project_id = sec.get("projectId")

    # данные по спецификации
    spec_items = (
        await db["spec_items"]
        .find(
            {
                "tenantId": tenant_id,
                "projectId": project_id,
                "sectionId": sec["_id"],
                "deleted": {"$ne": True},
            }
        )
        .sort([("pos", 1), ("createdAt", 1)])
        .to_list(100000)
    )

    # активная версия секции
    active_ver = (
        sec.get("activeVersion")
        or sec.get("version")
        or 1
    )

    def _item_version(it):
        return it.get("sectionVersion") or it.get("version") or 1

    visible_items = [
        it for it in spec_items if _item_version(it) == active_ver
    ]

    # только явные заголовки
    def _is_header_row(it: dict) -> bool:
        return bool(
            it.get("isHeader")
            or it.get("header")
            or it.get("rowType") == "header"
        )

    # разбиваем на блоки: заголовок + позиции под ним
    groups: list[dict] = []
    current: dict | None = None
    for it in visible_items:
        if _is_header_row(it):
            current = {"header": it, "items": []}
            groups.append(current)
        else:
            if not current:
                current = {"header": None, "items": []}
                groups.append(current)
            current["items"].append(it)

    flat_items = [it for g in groups for it in g["items"]]

    # отгрузка / выполнение / работы

    ship_items = await db["spec_ship_items"].find(
        {
            "tenantId": tenant_id,
            "projectId": project_id,
            "deleted": {"$ne": True},
        }
    ).to_list(100000)

    exec_items = await db["spec_exec_items"].find(
        {
            "tenantId": tenant_id,
            "projectId": project_id,
            "deleted": {"$ne": True},
        }
    ).to_list(100000)

    works = await db["spec_item_works"].find(
        {
            "tenantId": tenant_id,
            "projectId": project_id,
            "deleted": {"$ne": True},
        }
    ).to_list(100000)

    ship_by_item: dict[str, list[dict]] = defaultdict(list)
    for s in ship_items:
        sid = s.get("specItemId")
        if not sid:
            continue
        ship_by_item[str(sid)].append(s)

    exec_by_item: dict[str, list[dict]] = defaultdict(list)
    for e in exec_items:
        sid = e.get("specItemId")
        if not sid:
            continue
        exec_by_item[str(sid)].append(e)

    works_by_item: dict[str, list[dict]] = defaultdict(list)
    for w in works:
        iid = w.get("itemId")
        if not iid:
            continue
        works_by_item[str(iid)].append(w)

    # excel-книга

    wb = Workbook()
    ws = wb.active
    ws.title = "Свод"

    headers = [
        "№",
        "Наименование",
        "Ед. изм.",
        "Кол-во по спецификации",
        "Кол-во по отгрузке",
        "Кол-во фактическое",
    ]
    ws.append(headers)

    thin = Side(style="thin", color="000000")
    border_all = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="F2F2F2")
    work_fill = PatternFill("solid", fgColor="DDEBF7")

    # итоги по разделу
    total_work = 0.0
    total_mat = 0.0
    total_sum = 0.0

    row_idx = 2

    for grp in groups:
        header_item = grp.get("header")
        items_group = grp.get("items") or []

        # строка заголовка группы
        if header_item:
            name = header_item.get("name") or ""
            ws.cell(row=row_idx, column=1, value=name)
            ws.merge_cells(
                start_row=row_idx,
                start_column=1,
                end_row=row_idx,
                end_column=len(headers),
            )
            cell = ws.cell(row=row_idx, column=1)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(
                horizontal="left", vertical="center"
            )
            cell.fill = header_fill
            cell.border = border_all
            row_idx += 1

        # НУМЕРАЦИЯ С НУЛЯ ДЛЯ КАЖДОЙ ГРУППЫ
        pos_counter = 0

        for it in items_group:
            pos_counter += 1
            it_id_str = str(it["_id"])

            ship_list = ship_by_item.get(it_id_str, [])
            exec_list = exec_by_item.get(it_id_str, [])

            qty_spec = float(it.get("qty") or 0)
            qty_ship = sum(float(s.get("qty") or 0) for s in ship_list)
            qty_exec = sum(float(e.get("qty") or 0) for e in exec_list)

            ship_names = sorted(
                {
                    (s.get("name") or "").strip()
                    for s in ship_list
                    if s.get("name")
                }
            )
            base_name = it.get("name") or ""
            if ship_names:
                name = f"{base_name}\n(из отгрузки: {', '.join(ship_names)})"
            else:
                name = base_name

            ws.cell(row=row_idx, column=1, value=pos_counter)
            ws.cell(row=row_idx, column=2, value=name)
            ws.cell(row=row_idx, column=3, value=it.get("unit") or "")
            ws.cell(row=row_idx, column=4, value=qty_spec or None)
            ws.cell(row=row_idx, column=5, value=qty_ship or None)
            ws.cell(row=row_idx, column=6, value=qty_exec or None)

            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=row_idx, column=col)
                cell.border = border_all
                cell.alignment = Alignment(
                    wrap_text=True,
                    vertical="top",
                    horizontal="left" if col != 1 else "center",
                )
            
            # считаем итоги по разделу
            pw = float(it.get("price_work") or 0)
            pm = float(it.get("price_mat") or 0)
            work_cost = qty_spec * pw
            mat_cost = qty_spec * pm
            total_work += work_cost
            total_mat += mat_cost
            total_sum += work_cost + mat_cost

            row_idx += 1

            # работы под позицией без дублей
            seen_work_keys: set[tuple] = set()
            for w in works_by_item.get(it_id_str, []):
                key = (
                    (w.get("name") or "").strip(),
                    (w.get("unit") or "").strip(),
                    float(w.get("qty_plan") or 0),
                    float(w.get("qty_fact") or 0),
                )
                if key in seen_work_keys:
                    continue
                seen_work_keys.add(key)

                ws.cell(row=row_idx, column=1, value=None)
                ws.cell(
                    row=row_idx,
                    column=2,
                    value=f"Работа: {w.get('name') or ''}",
                )
                ws.cell(
                    row=row_idx,
                    column=3,
                    value=w.get("unit") or "",
                )
                ws.cell(
                    row=row_idx,
                    column=4,
                    value=w.get("qty_plan") or None,
                )
                ws.cell(
                    row=row_idx,
                    column=5,
                    value=None,
                )
                ws.cell(
                    row=row_idx,
                    column=6,
                    value=w.get("qty_fact") or None,
                )

                for col in range(1, len(headers) + 1):
                    cell = ws.cell(row=row_idx, column=col)
                    cell.border = border_all
                    cell.alignment = Alignment(
                        wrap_text=True,
                        vertical="top",
                        horizontal="left",
                    )
                    cell.fill = work_fill
                row_idx += 1

    # ширина колонок
    widths = {1: 5, 2: 70, 3: 10, 4: 20, 5: 20, 6: 20}
    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # формат чисел
    for row in ws.iter_rows(
        min_row=2, max_row=row_idx - 1, min_col=4, max_col=6
    ):
        for cell in row:
            if isinstance(cell.value, (int, float)):
                cell.number_format = "0.00"

    # шапка таблицы - жирная, по центру
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        cell.border = border_all

    # итоговая мини-таблица после основной

    row_idx += 1
    ws.cell(row=row_idx, column=2, value="Итого по разделу:")
    ws.cell(row=row_idx, column=3, value=sec.get("title") or "")
    ws.merge_cells(
        start_row=row_idx,
        start_column=3,
        end_row=row_idx,
        end_column=4,
    )
    row_idx += 1

    ws.cell(row=row_idx, column=2, value="Работы")
    ws.cell(row=row_idx, column=3, value=total_work)
    row_idx += 1

    ws.cell(row=row_idx, column=2, value="Материалы")
    ws.cell(row=row_idx, column=3, value=total_mat)
    row_idx += 1

    ws.cell(row=row_idx, column=2, value="Всего")
    ws.cell(row=row_idx, column=3, value=total_sum)

    for r in range(row_idx - 3, row_idx + 1):
        for c in range(2, 4):
            cell = ws.cell(row=r, column=c)
            cell.border = border_all
            cell.alignment = Alignment(
                horizontal="left", vertical="center"
            )
            if isinstance(cell.value, (int, float)) and c == 3:
                cell.number_format = "#,##0.00"

    # отдаём файл

    safe_title = (sec.get("title") or "summary").replace('"', "'").replace("\n", " ")
    ascii_title = re.sub(r"[^A-Za-z0-9_.-]+", "_", safe_title).strip("_") or "summary"
    filename = f"{ascii_title}.xlsx"

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)

    headers_resp = {"Content-Disposition": f'attachment; filename="{filename}"'}

    try:
        await log_project_action(
            db,
            user,
            project_id=project_id,
            action="spec.summary.export_excel",
            entity="spec.section",
            entity_id=str(sec["_id"]),
            message=f'Экспортирован свод по разделу «{safe_title}» в Excel',
            meta={"specSectionId": str(sec["_id"])},
        )
    except Exception:
        pass

    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers_resp,
    )
