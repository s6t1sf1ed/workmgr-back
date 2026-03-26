from fastapi import APIRouter, Depends, HTTPException, Query, Response
from bson import ObjectId
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment
from datetime import datetime, timezone, timedelta
import re
from collections import defaultdict
from .db import db
from . import auth
from .utils import to_jsonable
from .audit import log_user_action, make_diff
from .permissions import require_permission

# Время / парсинг

MSK = timezone(timedelta(hours=3))
UTC = timezone.utc

try:
    from dateutil.parser import isoparse as _parse_iso
except Exception:
    _parse_iso = None

DATE_ONLY_RX = re.compile(r"^\d{2}\.\d{2}\.\d{4}$|^\d{4}-\d{2}-\d{2}$")

def _is_date_only_str(x) -> bool:
    return isinstance(x, str) and bool(DATE_ONLY_RX.match(x.strip()))

def _day_bounds_msk(dt: datetime) -> tuple[datetime, datetime]:
    m = dt.astimezone(MSK)
    return (
        m.replace(hour=0, minute=0, second=0, microsecond=0),
        m.replace(hour=23, minute=59, second=59, microsecond=999_000),
    )

def _tenant_filter(user: dict) -> dict:

    """ Универсальный фильтр по tenantId: матчим и ObjectId, и строку. """

    tid = user.get("tenantId")
    if not tid:
        return {}
    t_oid = oid(tid)
    # строковое значение (и из самого user, и из ObjectId)
    t_str = str(t_oid) if isinstance(t_oid, ObjectId) else str(tid)
    return {
        "$or": [
            {"tenantId": t_oid},
            {"tenantId": t_str},
        ]
    }

# фильтр и сортировка
def _build_filter_and_sort(
    user,
    *,
    archived=None, q=None,
    person=None, project=None, telegram=None,
    startFrom=None, startTo=None,
    endFrom=None, endTo=None,
    hasPhoto=None,
    hoursMin=None, hoursMax=None,
    sort: str | None = "-start_time",
):
    filt: dict = {}
    and_conds: list[dict] = []

    # поддерживаем ObjectId и строку
    if user.get("tenantId"):
        t = oid(user["tenantId"])
        t_str = str(t)
        and_conds.append({
            "$or": [
                {"tenantId": t},
                {"tenantId": t_str},
                {"$expr": {"$eq": [{"$toString": "$tenantId"}, t_str]}},
            ]
        })

    # архив / не архив
    arch = _build_archived_filter(archived)
    if arch:
        and_conds.append(arch)

    if person:  and_conds.append(_match_person(person))
    if project: and_conds.append(_match_project(project))
    if telegram:
        and_conds.append({"$or":[
            {"telegram_id":{"$regex":telegram,"$options":"i"}},
            {"telegramId":{"$regex":telegram,"$options":"i"}},
        ]})
    if q:
        and_conds.append({"text_report":{"$regex":q,"$options":"i"}})

    # start_time диапазон
    raw_from, raw_to = startFrom, startTo
    dt_from = _as_dt_loose(raw_from) if raw_from else None
    dt_to   = _as_dt_loose(raw_to)   if raw_to   else None

    if dt_from and _is_date_only_str(raw_from):
        dt_from, _ = _day_bounds_msk(dt_from)
    if dt_to and _is_date_only_str(raw_to):
        _, dt_to = _day_bounds_msk(dt_to)

    cond_start: dict = {}
    if dt_from:
        cond_start["$gte"] = dt_from.astimezone(UTC)
    if dt_to:
        cond_start["$lte"] = dt_to.astimezone(UTC)
    if cond_start:
        and_conds.append({"start_time": cond_start})

    # end_time диапазон
    raw_efrom, raw_eto = endFrom, endTo
    et_from = _as_dt_loose(raw_efrom) if raw_efrom else None
    et_to   = _as_dt_loose(raw_eto)   if raw_eto   else None
    if et_from and _is_date_only_str(raw_efrom):
        et_from, _ = _day_bounds_msk(et_from)
    if et_to and _is_date_only_str(raw_eto):
        _, et_to = _day_bounds_msk(et_to)

    cond_end: dict = {}
    if et_from:
        cond_end["$gte"] = et_from.astimezone(UTC)
    if et_to:
        cond_end["$lte"] = et_to.astimezone(UTC)
    if cond_end:
        and_conds.append({"end_time": cond_end})

    if hasPhoto is not None:
        if int(hasPhoto) == 1:
            and_conds.append({"photo_link":{"$exists":True,"$ne":""}})
        else:
            and_conds.append({"$or":[{"photo_link":""},{"photo_link":{"$exists":False}}]})

    exprs = []
    if hoursMin is not None or hoursMax is not None:
        hours_expr = {
            "$divide": [
                {
                    "$subtract": [
                        {"$ifNull": ["$end_time", "$start_time"]},
                        "$start_time",
                    ]
                },
                3600000,
            ]
        }
        if hoursMin is not None:
            exprs.append({"$gte": [hours_expr, float(hoursMin)]})
        if hoursMax is not None:
            exprs.append({"$lte": [hours_expr, float(hoursMax)]})


    if and_conds: filt.setdefault("$and", []).extend(and_conds)
    if exprs:     filt.setdefault("$expr", {"$and": exprs})

    s = [("start_time",-1), ("_id",-1)]
    if sort:
        desc = sort.startswith("-")
        fld  = sort[1:] if desc else sort
        s = [(fld, -1 if desc else 1), ("_id",-1)]

    return filt, s

# генератор с подгрузкой ссылок
async def _iter_reports_with_refs(filt, sort):
    pcache: dict[str, dict] = {}
    prcache: dict[str, dict] = {}
    async for r in db.reports.find(filt).sort(sort):
        pid = r.get("person_id") or r.get("personId") or r.get("user_id")
        pid_key = str(pid) if isinstance(pid, ObjectId) else str(pid or "")
        if pid and pid_key not in pcache:
            pcache[pid_key] = await _load_person_by_any_id(pid)
        r["_person"] = pcache.get(pid_key, {})

        prid = r.get("project_id") or r.get("projectId")
        prid_key = str(prid) if isinstance(prid, ObjectId) else str(prid or "")
        if prid and prid_key not in prcache:
            prcache[prid_key] = await _load_project_by_any_id(prid)
        r["_project"] = prcache.get(prid_key, {})

        yield r

def _try_strptime(s: str) -> datetime | None:
    s = s.strip()
    for fmt in ("%d.%m.%Y, %H:%M", "%d.%m.%Y %H:%M"):
        try:
            return datetime.strptime(s, fmt).replace(tzinfo=MSK)
        except Exception:
            pass
    return None

def _as_dt_loose(x) -> datetime | None:
    if x is None or x == "":
        return None
    if isinstance(x, datetime):
        return x.replace(tzinfo=UTC) if x.tzinfo is None else x
    if isinstance(x, (int, float)):
        if x > 10_000_000_000: x = x / 1000.0
        return datetime.fromtimestamp(float(x), tz=UTC)
    if isinstance(x, str):
        s = x.strip()
        if not s: return None
        dt_local = _try_strptime(s)
        if dt_local: return dt_local
        if _parse_iso:
            try:
                dt = _parse_iso(s)
                if dt.tzinfo is None: dt = dt.replace(tzinfo=MSK)
                return dt
            except Exception:
                pass
        try:
            if s.endswith("Z"): s = s[:-1] + "+00:00"
            dt = datetime.fromisoformat(s)
            if dt.tzinfo is None: dt = dt.replace(tzinfo=MSK)
            return dt
        except Exception:
            return None
    return None

def _to_iso_utc(x) -> str | None:
    if x is None or x == "": return None
    if isinstance(x, datetime):
        dt = x if x.tzinfo else x.replace(tzinfo=UTC)
        return dt.astimezone(UTC).isoformat().replace("+00:00","Z")
    dt = _as_dt_loose(x)
    if not dt: return None
    if dt.tzinfo is None: dt = dt.replace(tzinfo=MSK)
    return dt.astimezone(UTC).isoformat().replace("+00:00","Z")

def _as_utc_for_storage(x) -> datetime | None:
    dt = _as_dt_loose(x)
    if not dt: return None
    return dt.astimezone(UTC)

def _to_msk(x) -> datetime | None:
    dt = _as_dt_loose(x)
    if not dt: return None
    return dt.astimezone(MSK)

def _fmt_date_time_range(start, end) -> str:
    s = _to_msk(start)
    if not s: return ""
    e = _to_msk(end) if end else None
    return f"{s.strftime('%d.%m.%Y')}\n{s.strftime('%H:%M')}-{e.strftime('%H:%M') if e else '—'}"

def _fmt_duration(start, end) -> str:
    s = _to_msk(start); e = _to_msk(end)
    if not s or not e: return ""
    minutes = int((e - s).total_seconds() // 60)
    return f"{minutes // 60}:{minutes % 60:02d}"

WEEKDAYS_RU = [
    "понедельник",
    "вторник",
    "среда",
    "четверг",
    "пятница",
    "суббота",
    "воскресенье",
]

def _duration_minutes(start, end) -> int:
    s = _to_msk(start); e = _to_msk(end)
    if not s or not e:
        return 0
    if e < s:
        return 0
    return int((e - s).total_seconds() // 60)

def _weekday_name_ru_by_start(start) -> str:
    """Имя дня недели по дате начала (в МСК)"""
    dt = _to_msk(start)
    if not dt:
        return ""
    idx = dt.weekday()
    try:
        return WEEKDAYS_RU[idx]
    except Exception:
        return ""

# Утилиты / нормализация

router = APIRouter(prefix="/api/reports", tags=["reports"])

def oid(v):
    if isinstance(v, ObjectId): return v
    try: return ObjectId(v)
    except Exception: return v

def _norm_id(x): return str(x) if isinstance(x, ObjectId) else x

def _fio_from_doc(p: dict | None) -> str:
    if not p: return ""
    ln = (p.get("lastName") or p.get("surname") or p.get("last_name") or "").strip()
    fn = (p.get("firstName") or p.get("name") or p.get("first_name") or "").strip()
    if ln or fn: return " ".join(x for x in (ln, fn) if x).strip()
    return (p.get("name") or p.get("email") or p.get("telegramId") or p.get("telegram_id") or "").strip()

def _tg_from_doc(p: dict | None) -> str:
    if not p: return ""
    v = p.get("telegramId") or p.get("telegram_id") or ""
    return "" if v is None else str(v)

async def _load_person_by_any_id(person_any) -> dict:
    if not person_any: return {}
    pid = oid(person_any)
    if isinstance(pid, ObjectId):
        for coll in ("person","persons","users"):
            p = await db[coll].find_one({"_id": pid})
            if p: return p
    return {}

async def _load_project_by_any_id(project_any) -> dict:
    if not project_any: return {}
    pid = oid(project_any)
    if isinstance(pid, ObjectId):
        for coll in ("projects","project"):
            p = await db[coll].find_one({"_id": pid})
            if p: return p
    return {}

def normalize_report(d: dict) -> dict:
    if not d: return {}
    out = dict(d)
    out["start_time"] = _to_iso_utc(d.get("start_time"))
    out["end_time"]   = _to_iso_utc(d.get("end_time"))
    out["_id"] = _norm_id(out.get("_id"))
    for k in ("person_id","personId","project_id","projectId","tenantId","user_id","company","companyId","company_id"):
        if k in out: out[k] = _norm_id(out[k])
    p = d.get("_person") or {}
    pr = d.get("_project") or {}
    out["person"]  = {"id": _norm_id(d.get("person_id") or d.get("personId") or d.get("user_id")),
                      "name": _fio_from_doc(p), "telegram_id": _tg_from_doc(p)}
    out["project"] = {"id": _norm_id(d.get("project_id") or d.get("projectId")), "name": pr.get("name","") or ""}
    st, et = d.get("start_time"), d.get("end_time")
    out["timeRangeHuman"] = _fmt_date_time_range(st, et)
    out["workHoursHuman"] = _fmt_duration(st, et)
    out.pop("_person", None); out.pop("_project", None)
    return to_jsonable(out)

def _report_title(d: dict) -> str:

    """ Человекочитаемый заголовок отчёта для логов: "Фамилия Имя / Проект" """
    p = d.get("_person") or {}
    pr = d.get("_project") or {}
    fio = _fio_from_doc(p)
    proj = (pr.get("name") or "").strip()
    if fio and proj:
        return f"{fio} / {proj}"
    return fio or proj or "Отчёт"


def _report_meta(d: dict) -> dict:

    """ Компактные метаданные для логов: id сотрудника/проекта и время """
    pid = d.get("person_id") or d.get("personId") or d.get("user_id")
    prj = d.get("project_id") or d.get("projectId")
    return {
        "personId": _norm_id(pid) if pid else None,
        "projectId": _norm_id(prj) if prj else None,
        "start_time": _to_iso_utc(d.get("start_time")),
        "end_time": _to_iso_utc(d.get("end_time")),
    }

def _build_archived_filter(archived: int | None):
    if archived is None: return {}
    if int(archived) == 1: return {"archived": True}
    return {"$or":[{"archived": False}, {"archived": {"$exists": False}}]}

def _match_person(pid_str: str):
    oid_val = oid(pid_str); str_val = str(pid_str)
    return {"$or":[
        {"person_id": oid_val}, {"personId": oid_val}, {"user_id": oid_val},
        {"person_id": str_val}, {"personId": str_val}, {"user_id": str_val},
        {"$expr":{"$eq":[{"$toString":"$person_id"}, str_val]}},
        {"$expr":{"$eq":[{"$toString":"$personId"}, str_val]}},
        {"$expr":{"$eq":[{"$toString":"$user_id"}, str_val]}},
    ]}

def _match_project(prj_str: str):
    oid_val = oid(prj_str); str_val = str(prj_str)
    return {"$or":[
        {"project_id": oid_val}, {"projectId": oid_val},
        {"project_id": str_val}, {"projectId": str_val},
        {"$expr":{"$eq":[{"$toString":"$project_id"}, str_val]}},
        {"$expr":{"$eq":[{"$toString":"$projectId"}, str_val]}},
    ]}

def _set_dt_if_present(payload: dict, field: str):
    if field in payload:
        v = payload.get(field)
        dt = _as_utc_for_storage(v)
        if dt is not None: payload[field] = dt
        elif v == "" or v is None: payload.pop(field, None)

# Роуты

@router.get("")
async def list_reports(
    user=Depends(auth.get_current_user),
    page: int = 1,
    limit: int = 50,
    sort: str = "-start_time",
    archived: int | None = None,
    q: str | None = None,
    person: str | None = Query(None, alias="personId"),
    project: str | None = Query(None, alias="projectId"),
    telegram: str | None = None,
    startFrom: str | None = None,
    startTo: str | None = None,
    endFrom: str | None = None,
    endTo: str | None = None,
    hasPhoto: int | None = None,
    hoursMin: float | None = None,
    hoursMax: float | None = None,
):
    require_permission(user, "reports.view")
    filt, s = _build_filter_and_sort(
        user,
        archived=archived, q=q, person=person, project=project, telegram=telegram,
        startFrom=startFrom, startTo=startTo, endFrom=endFrom, endTo=endTo,
        hasPhoto=hasPhoto, hoursMin=hoursMin, hoursMax=hoursMax, sort=sort
    )

    total = await db.reports.count_documents(filt)

    MAX_ALL = 5000
    effective_limit = min(total, MAX_ALL) if page == 1 else min(limit, MAX_ALL)

    items: list[dict] = []
    i = 0
    async for r in _iter_reports_with_refs(filt, s):
        if i < (page - 1) * limit:
            i += 1
            continue
        if len(items) >= effective_limit:
            break
        items.append(normalize_report(r))
        i += 1

    return to_jsonable({"items": items, "page": page, "limit": effective_limit, "total": total})

@router.get("/{rid}")
async def get_one(rid: str, user=Depends(auth.get_current_user)):
    require_permission(user, "reports.view")
    # ищем по id + tenantId (строка или ObjectId)
    base = {"_id": oid(rid)}
    if user.get("tenantId"):
        t = oid(user["tenantId"])
        t_str = str(t)
        base["$or"] = [
            {"tenantId": t},
            {"tenantId": t_str},
            {"$expr": {"$eq": [{"$toString": "$tenantId"}, t_str]}},
        ]

    d = await db.reports.find_one(base)
    if not d:
        raise HTTPException(404, "not found")

    pid = d.get("person_id") or d.get("personId") or d.get("user_id")
    d["_person"] = await _load_person_by_any_id(pid)
    pr = d.get("project_id") or d.get("projectId")
    d["_project"] = await _load_project_by_any_id(pr)
    return to_jsonable(normalize_report(d))

@router.post("")
async def create(payload: dict, user=Depends(auth.get_current_user)):
    require_permission(user, "reports.create")
    _set_dt_if_present(payload, "start_time")
    _set_dt_if_present(payload, "end_time")
    payload["tenantId"] = oid(user["tenantId"]) if user.get("tenantId") else None
    payload["createdAt"] = datetime.now(tz=UTC)
    payload["updatedAt"] = datetime.now(tz=UTC)
    res = await db.reports.insert_one(payload)
    d = await db.reports.find_one({"_id": res.inserted_id})
    pid = d.get("person_id") or d.get("personId") or d.get("user_id")
    d["_person"] = await _load_person_by_any_id(pid)
    pr = d.get("project_id") or d.get("projectId")
    d["_project"] = await _load_project_by_any_id(pr)

    try:
        await log_user_action(
            db,
            user,
            action="report.create",
            entity="report",
            entity_id=str(d["_id"]),
            message=f'Создан отчёт «{_report_title(d)}»',
            meta=_report_meta(d),
        )
    except Exception:
        pass

    return to_jsonable(normalize_report(d))

@router.patch("/{rid}")
async def update(rid: str, payload: dict, user=Depends(auth.get_current_user)):
    require_permission(user, "reports.edit")
    _set_dt_if_present(payload, "start_time")
    _set_dt_if_present(payload, "end_time")
    payload["updatedAt"] = datetime.now(tz=UTC)

    # ДО изменений
    base = {"_id": oid(rid)}
    if user.get("tenantId"):
        t = oid(user["tenantId"])
        t_str = str(t)
        base["$or"] = [
            {"tenantId": t},
            {"tenantId": t_str},
            {"$expr": {"$eq": [{"$toString": "$tenantId"}, t_str]}},
        ]

    before = await db.reports.find_one(base)
    if not before:
        raise HTTPException(404, "not found")

    await db.reports.update_one({"_id": before["_id"]}, {"$set": payload})
    d = await db.reports.find_one({"_id": before["_id"]})

    pid = d.get("person_id") or d.get("personId") or d.get("user_id")
    d["_person"] = await _load_person_by_any_id(pid)
    pr = d.get("project_id") or d.get("projectId")
    d["_project"] = await _load_project_by_any_id(pr)

    try:
        diff = make_diff(before, d)
        if diff:
            await log_user_action(
                db,
                user,
                action="report.update",
                entity="report",
                entity_id=str(d["_id"]),
                message=f'Обновлён отчёт «{_report_title(d)}»',
                diff=diff,
                meta=_report_meta(d),
            )
    except Exception:
        pass

    return to_jsonable(normalize_report(d))

@router.delete("/{rid}")
async def delete(rid: str, user=Depends(auth.get_current_user)):
    require_permission(user, "reports.delete")
    base = {"_id": oid(rid)}
    if user.get("tenantId"):
        t = oid(user["tenantId"])
        t_str = str(t)
        base["$or"] = [
            {"tenantId": t},
            {"tenantId": t_str},
            {"$expr": {"$eq": [{"$toString": "$tenantId"}, t_str]}},
        ]

    d = await db.reports.find_one(base)
    if not d:
        raise HTTPException(404, "not found")

    pid = d.get("person_id") or d.get("personId") or d.get("user_id")
    d["_person"] = await _load_person_by_any_id(pid)
    pr = d.get("project_id") or d.get("projectId")
    d["_project"] = await _load_project_by_any_id(pr)

    await db.reports.delete_one({"_id": d["_id"]})
    try:
        await log_user_action(
            db,
            user,
            action="report.delete",
            entity="report",
            entity_id=str(rid),
            message=f'Удалён отчёт «{_report_title(d)}»',
            meta=_report_meta(d),
        )
    except Exception:
        pass

    return {"ok": True}

@router.get("/export/xlsx")
async def export_xlsx(
    user=Depends(auth.get_current_user),
    archived: int | None = None,
    q: str | None = None,
    person: str | None = Query(None, alias="personId"),
    project: str | None = Query(None, alias="projectId"),
    telegram: str | None = None,
    startFrom: str | None = None,
    startTo: str | None = None,
    endFrom: str | None = None,
    endTo: str | None = None,
    hasPhoto: int | None = None,
    hoursMin: float | None = None,
    hoursMax: float | None = None,
    sort: str | None = "-start_time",
):
    require_permission(user, "reports.view")
    filt, s = _build_filter_and_sort(
        user,
        archived=archived, q=q, person=person, project=project, telegram=telegram,
        startFrom=startFrom, startTo=startTo, endFrom=endFrom, endTo=endTo,
        hasPhoto=hasPhoto, hoursMin=hoursMin, hoursMax=hoursMax, sort=sort
    )

    rows: list[dict] = []
    async for r in _iter_reports_with_refs(filt, s):
        rows.append(r)

    wb = Workbook()
    ws = wb.active
    ws.title = "Отчёты"
    headers = ["Telegram ID","Сотрудник","Проект","Дата / время","День недели","Часы","Текст отчёта","Фото", "Лог сессии"]
    for c, v in enumerate(headers, start=1):
        ws.cell(row=1, column=c, value=v)

    row = 2
    for d in rows:
        p = d.get("_person") or {}
        prj = d.get("_project") or {}
        values = [
            _tg_from_doc(p),
            _fio_from_doc(p),
            prj.get("name","") or "",
            _fmt_date_time_range(d.get("start_time"), d.get("end_time")),
            _weekday_name_ru_by_start(d.get("start_time")),
            _fmt_duration(d.get("start_time"), d.get("end_time")),
            d.get("text_report") or "",
            d.get("photo_link") or "",
            d.get("session_log") or "",
        ]
        for col_idx, val in enumerate(values, start=1):
            ws.cell(row=row, column=col_idx, value=val)
        row += 1

    out = BytesIO(); wb.save(out); out.seek(0)
    headers = {
        "Content-Disposition": 'attachment; filename="reports.xlsx"',
        "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    }

    try:
        await log_user_action(
            db,
            user,
            action="report.export.xlsx",
            entity="report",
            entity_id=None,
            message="Экспорт отчётов в Excel",
            meta={
                "rows": len(rows),
                "filter": filt,
            },
        )
    except Exception:
        pass

    return Response(content=out.getvalue(), headers=headers, media_type=headers["Content-Type"])

@router.get("/timesheet/xlsx")
async def export_timesheet_xlsx(
    user=Depends(auth.get_current_user),
    person: str = Query(..., alias="personId"),
    month: str = Query(..., description="YYYY-MM"),
):
    require_permission(user, "timesheet.view")

    # парсим месяц и считаем границы в МСК
    try:
        year_str, month_str = month.split("-")
        year = int(year_str)
        m = int(month_str)
        if not (1 <= m <= 12):
            raise ValueError
    except Exception:
        raise HTTPException(400, "month must be in format YYYY-MM")

    month_start_msk = datetime(year, m, 1, tzinfo=MSK)
    if m == 12:
        next_start_msk = datetime(year + 1, 1, 1, tzinfo=MSK)
    else:
        next_start_msk = datetime(year, m + 1, 1, tzinfo=MSK)

    # достаём все отчёты сотрудника за месяц
    start_iso = month_start_msk.isoformat()
    end_iso = next_start_msk.isoformat()

    filt, sort_spec = _build_filter_and_sort(
        user,
        archived=0,
        person=person,
        startFrom=start_iso,
        startTo=end_iso,
        sort="start_time",
    )

    reports: list[dict] = []
    async for r in _iter_reports_with_refs(filt, sort_spec):
        reports.append(r)

    # базовая инфа по сотруднику
    person_doc: dict = {}
    if reports:
        person_doc = reports[0].get("_person") or {}
        if not person_doc:
            pid_any = (
                reports[0].get("person_id")
                or reports[0].get("personId")
                or reports[0].get("user_id")
            )
            person_doc = await _load_person_by_any_id(pid_any)
    else:
        # если за месяц отчётов нет
        person_doc = await _load_person_by_any_id(person)

    fio_value = _fio_from_doc(person_doc)
    tg_value = _tg_from_doc(person_doc)

    # группируем отчёты по дате 
    by_date: dict[datetime.date, list[dict]] = defaultdict(list)
    for r in reports:
        st = r.get("start_time")
        dt_msk = _to_msk(st)
        if not dt_msk:
            continue
        by_date[dt_msk.date()].append(r)

    # Excel в формате шаблона
    wb = Workbook()
    ws = wb.active
    ws.title = "Табель"

    # Заголовки
    headers = [
        "Telegram ID",
        "Сотрудник",  
        "Проект",
        "Дата / время",
        "День недели",
        "Норма часов",
        "Часы", # ([h]:mm)
        "Выработка", # (=G*24)
        "Оплата", # (=H*500)
        "Бонус",
        "Косяки",
        "Комментарий",
    ]
    for c, v in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=c, value=v)
        cell.font = cell.font.copy(bold=True)

    # Ширины колонок
    widths = {
        "A": 13.0,
        "B": 18.86,
        "C": 26.43,
        "D": 11.71,
        "E": 11.29,
        "F": 11.86,
        "G": 10.57,
        "H": 11.86,
        "I": 11.86,
        "J": 10.43,
        "K": 11.57,
        "L": 26.43,
        "M": 2.43,
        "N": 15.86,
        "O": 13.43,
        "P": 3.86,
        "Q": 11.57,
        "R": 3.71,
    }
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    
    # стили
    gray_fill = PatternFill("solid", fgColor="D9D9D9")   # выходные
    orange_fill = PatternFill("solid", fgColor="FFF2CC") # J/K
    thin_side = Side(style="thin", color="000000")
    thin_border = Border(top=thin_side, bottom=thin_side, left=thin_side, right=thin_side)

    row_idx = 2
    first_data_row = 2

    # для половины месяца (1–15)
    half_worked_days = 0         # отработано дней
    half_norm_days = 0           # рабочих дней 1-15
    half_total_minutes = 0       # отработано минут 1-15
    first_half_start_row = None  # строка для 15-го числа
    first_half_end_row = None    # строка для 1-го числа

    # от последнего дня месяца к первому
    cur = next_start_msk - timedelta(days=1)
    while cur >= month_start_msk:
        d_date = cur.date()
        day_reports = by_date.get(d_date, [])

        weekday_idx = cur.weekday()
        weekday_name = WEEKDAYS_RU[weekday_idx]
        norm_hours = 8 if weekday_idx < 5 else 0

        # проекты за день (если несколько - через запятую)
        proj_names: set[str] = set()
        for r in day_reports:
            pr = r.get("_project") or {}
            name = (pr.get("name") or "").strip()
            if name:
                proj_names.add(name)
        project_text = ", ".join(sorted(proj_names))

        # суммарные минуты за день
        total_minutes = sum(
            _duration_minutes(r.get("start_time"), r.get("end_time"))
            for r in day_reports
        )
        excel_hours_value = total_minutes / (24 * 60) if total_minutes > 0 else 0

        # дата
        current_row = row_idx
        date_cell_value = datetime(cur.year, cur.month, cur.day)

        # значения
        ws.cell(row=current_row, column=1, value=tg_value)
        ws.cell(row=current_row, column=2, value=fio_value)
        ws.cell(row=current_row, column=3, value=project_text)

        c_date = ws.cell(row=current_row, column=4, value=date_cell_value)
        c_date.number_format = "dd.mm.yyyy"

        ws.cell(row=current_row, column=5, value=weekday_name)
        ws.cell(row=current_row, column=6, value=norm_hours)

        c_hours = ws.cell(row=current_row, column=7, value=excel_hours_value)
        c_hours.number_format = "[h]:mm"

        # выработка = G*24
        c_work = ws.cell(row=current_row, column=8, value=f"=G{current_row}*24")
        c_work.number_format = "0.00"

        # оплата = H*500
        c_pay = ws.cell(row=current_row, column=9, value=f"=H{current_row}*$N$22")
        c_pay.number_format = "0.00"

        # бонус и косяки
        c_bonus = ws.cell(row=current_row, column=10, value=None)
        c_bonus.number_format = "0.00"

        c_mist = ws.cell(row=current_row, column=11, value=None)
        c_mist.number_format = "0.00"

        # комментарий "Выходной"  субб./воск.
        comment_text = "Выходной" if weekday_idx >= 5 else ""
        ws.cell(row=current_row, column=12, value=comment_text)

        for col in range(1, 13):
            cell = ws.cell(row=current_row, column=col)
            if weekday_idx >= 5:
                cell.fill = gray_fill
            elif col in (10, 11):
                cell.fill = orange_fill

        if cur.day <= 15:
            if total_minutes > 0:
                half_worked_days += 1
            if norm_hours > 0:
                half_norm_days += 1
            half_total_minutes += total_minutes

            if cur.day == 15:
                first_half_start_row = current_row
            if cur.day == 1:
                first_half_end_row = current_row

        row_idx += 1
        cur -= timedelta(days=1)

    last_data_row = row_idx - 1

    # итоги по таблице
    total_row = None
    if last_data_row >= first_data_row:
        total_row = row_idx

        ws.cell(row=total_row, column=5, value="Итого:")

        ws.cell(
            row=total_row,
            column=6,
            value=f"=SUM(F{first_data_row}:F{last_data_row})", # Норма часов
        )

        total_hours_cell = ws.cell(
            row=total_row,
            column=7,
            value=f"=SUM(G{first_data_row}:G{last_data_row})", # Часы
        )
        total_hours_cell.number_format = "[h]:mm"

        total_work_cell = ws.cell(
            row=total_row,
            column=8,
            value=f"=SUM(H{first_data_row}:H{last_data_row})", # Выработка
        )
        total_work_cell.number_format = "0.00"

        total_pay_cell = ws.cell(
            row=total_row,
            column=9,
            value=f"=SUM(I{first_data_row}:I{last_data_row})", # Оплата
        )
        total_pay_cell.number_format = "0.00"

        total_bonus_cell = ws.cell(
            row=total_row,
            column=10,
            value=f"=SUM(J{first_data_row}:J{last_data_row})", # Бонус
        )
        total_bonus_cell.number_format = "0.00"

        total_mist_cell = ws.cell(
            row=total_row,
            column=11,
            value=f"=SUM(K{first_data_row}:K{last_data_row})", # Косяки
        )
        total_mist_cell.number_format = "0.00"

        # вспомогательные значения для половины месяца
        half_norm_hours_total = half_norm_days * 8
        half_hours_value = half_total_minutes / (24 * 60) if half_total_minutes > 0 else 0

        # строка диапазона 1-15 для суммы оплаты
        if first_half_start_row is not None and first_half_end_row is not None:
            half_pay_formula = f"=SUM(I{first_half_start_row}:I{first_half_end_row})"
        else:
            half_pay_formula = "0"

        # АВАНС
        ws["N2"] = "Аванс"
        ws.merge_cells("N2:Q2")
        ws["N2"].alignment = Alignment(horizontal="center")

        ws["N3"] = "Отработано"

        # диапазон строк для 1-15 числа
        if first_half_start_row is not None and first_half_end_row is not None:
            first_r = min(first_half_start_row, first_half_end_row)
            last_r = max(first_half_start_row, first_half_end_row)
            g_range = f"G{first_r}:G{last_r}"
            f_range = f"F{first_r}:F{last_r}"
        else:
            g_range = None
            f_range = None

        ws["N4"] = "Дней"
        if g_range and f_range:
            ws["O4"] = f'=COUNTIF({g_range},">0")' # отработано дней за 1-15
            ws["Q4"] = f'=COUNTIF({f_range},">0")' # рабочих дней за 1-15
        else:
            ws["O4"] = 0
            ws["Q4"] = 0
        ws["P4"] = "Из"

        ws["N5"] = "Часов"
        c_o5 = ws["O5"]
        if g_range and f_range:
            c_o5.value = f"=SUM({g_range})" # отработано часов за 1-15
            ws["Q5"] = f"=SUM({f_range})" # нормочасы за 1-15
        else:
            c_o5.value = 0
            ws["Q5"] = 0
        c_o5.number_format = "[h]:mm"
        ws["P5"] = "Из"

        ws["N6"] = "Итого:"
        c_o6 = ws["O6"]
        c_o6.value = half_pay_formula # сумма оплаты за 1-15
        c_o6.number_format = "0.00"

        for r in range(2, 7):
            for c in range(14, 18):
                ws.cell(row=r, column=c).border = thin_border

        # МЕСЯЦ
        ws["N8"] = "Месяц"
        ws.merge_cells("N8:Q8")
        ws["N8"].alignment = Alignment(horizontal="center")

        ws["N9"] = "Отработано"

        ws["N10"] = "Дней"
        ws["O10"] = f"=COUNTIF(G{first_data_row}:G{last_data_row},\">0\")"
        ws["P10"] = "Из"
        ws["Q10"] = f"=COUNTIF(F{first_data_row}:F{last_data_row},\">0\")"

        ws["N11"] = "Часов"
        c_o11 = ws["O11"]
        c_o11.value = f"=SUM(G{first_data_row}:G{last_data_row})"
        c_o11.number_format = "[h]:mm"
        ws["P11"] = "Из"
        ws["Q11"] = f"=SUM(F{first_data_row}:F{last_data_row})"

        for r in range(8, 12):
            for c in range(14, 18):
                ws.cell(row=r, column=c).border = thin_border

        # К ПОЛУЧЕНИЮ
        ws["N14"] = "К получению"
        ws.merge_cells("N14:Q14")
        ws["N14"].alignment = Alignment(horizontal="center")

        ws["N15"] = "Выработка:"
        c_o15 = ws["O15"]
        c_o15.value = f"=I{total_row}"
        c_o15.number_format = "0.00"

        ws["N16"] = "Бонус:"
        c_o16 = ws["O16"]
        c_o16.value = f"=J{total_row}"
        c_o16.number_format = "0.00"

        ws["N17"] = "Косяки:"
        c_o17 = ws["O17"]
        c_o17.value = f"=K{total_row}"
        c_o17.number_format = "0.00"

        ws["N18"] = "Аванс:"
        c_o18 = ws["O18"]
        c_o18.value = c_o6.value
        c_o18.number_format = "0.00"

        ws["N19"] = "Итого:"
        c_o19 = ws["O19"]
        c_o19.value = "=SUM(O15:O17)"
        c_o19.number_format = "0.00"

        ws["N21"] = "Ставка за час"
        rate_cell = ws["N22"]
        rate_cell.value = 500
        rate_cell.number_format = "0.00"

        for r in (21, 22):
            ws.cell(row=r, column=14).border = thin_border

        for r in range(14, 20):
            for c in range(14, 18):
                ws.cell(row=r, column=c).border = thin_border

        table_last_row = total_row
        for r in range(1, table_last_row + 1):
            for c in range(1, 12 + 1):
                cell = ws.cell(row=r, column=c)
                cell.border = thin_border

        

    ws.freeze_panes = "A2"

    out = BytesIO()
    wb.save(out)
    out.seek(0)

    headers_resp = {
        "Content-Disposition": 'attachment; filename="timesheet.xlsx"',
        "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    }

    try:
        await log_user_action(
            db,
            user,
            action="report.timesheet.xlsx",
            entity="report",
            entity_id=None,
            message=f"Экспорт табеля за месяц {month}",
            meta={
                "personId": person,
                "month": month,
            },
        )
    except Exception:
        pass

    return Response(
        content=out.getvalue(),
        headers=headers_resp,
        media_type=headers_resp["Content-Type"],
    )